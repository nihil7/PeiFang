# B05B_build_frame.py
# 功能：
# 1) 读取 output 下最新 tasks_prepared_*.json
# 2) 选择时间窗 / 机台排序与隐藏 / lanes 分配（含结束日）
# 3) 自动列宽（按 max_len；跨天任务不参与；可硬排除某些 machine）
# 4) 生成框架 xlsx（块边框：无 lane 内横线；只保留日期竖线 + 机台块上下粗线）
# 5) 输出 layout.json（给 B06 用）

import os
import sys
import json
from datetime import datetime, timedelta, date
from typing import Dict, List, Optional, Tuple
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from peifang_core.common import ROOT_DIR
from peifang_core.schedule_web import render_schedule_html


# =========================
# 配置区（只改这里）
# =========================
BASE_DIR = str(ROOT_DIR)
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

AUTO_LATEST = True
PREPARED_JSON = os.path.join(OUTPUT_DIR, "tasks_prepared_YYYYMMDD_HHMMSS.json")

TITLE = "机台×日期排程（框架）"

# 时间窗：THIS_WEEK / ROLLING / CUSTOM
DATE_PRESET = "ROLLING"
PAST_DAYS = 2
FUTURE_DAYS = 10
START_DATE = None  # "2026-01-05" 或 None
END_DATE = None    # "2026-01-20" 或 None
MAX_DAYS = 31

# 机台排序/显示
MACHINE_ORDER: List[str] = ["35机", "4#65机", "1#机", "2#机", "生产日记"]
INCLUDE_OTHER_MACHINES = False
HIDE_EMPTY_MACHINES = True
HIDE_MACHINES: List[str] = []

# 行高与列宽
MACHINE_COL_WIDTH = 14
HEADER_ROW_HEIGHT = 26
BASE_LANE_ROW_HEIGHT = 44

# Excel 网格线显示
SHOW_EXCEL_GRIDLINES = False

# ========= 参考文件边框参数对齐（关键） =========
# 日期竖向分割线（参考文件：thin + 00E5E7EB）
DATE_VLINE_STYLE = "thin"
DATE_VLINE_COLOR = "00E5E7EB"

# 机台块上下分隔线（参考文件：medium + 00CBD5E1）
APPLY_MACHINE_SEPARATORS = True
MACHINE_SEP_STYLE = "medium"
MACHINE_SEP_COLOR = "00CBD5E1"
MACHINE_SEP_TOP = True
MACHINE_SEP_BOTTOM = True

# A列与日期区分割线（参考文件：medium + 00E5E7EB）
MACHINE_DATE_VLINE_STYLE = "medium"
MACHINE_DATE_VLINE_COLOR = "00E5E7EB"
# =============================================

# 自动列宽（按 prepared 的 max_len）
AUTO_DATE_COL_WIDTH = True
DATE_COL_WIDTH_MIN = 16
DATE_COL_WIDTH_MAX = 50
COL_PADDING_CHARS = 4
COL_CHAR_FACTOR = 1.2

# 列宽统计硬排除的机台（这些任务不影响列宽）
EXCLUDE_WIDTH_MACHINES: List[str] = ["生产日记"]

PRINT_FIRST_N = 20
# =========================


def pick_latest_prepared(output_dir: str) -> Optional[str]:
    if not os.path.isdir(output_dir):
        return None
    files = [f for f in os.listdir(output_dir) if f.startswith("tasks_prepared_") and f.endswith(".json")]
    if not files:
        return None
    files.sort(key=lambda fn: os.path.getmtime(os.path.join(output_dir, fn)), reverse=True)
    return os.path.join(output_dir, files[0])


def parse_ymd(x: Optional[str]) -> Optional[date]:
    if not x:
        return None
    return datetime.strptime(x, "%Y-%m-%d").date()


def weekday_cn(d: date) -> str:
    names = ["一", "二", "三", "四", "五", "六", "日"]
    return f"周{names[d.weekday()]}"


def build_dates(start: date, end: date) -> List[date]:
    out = []
    d = start
    while d <= end:
        out.append(d)
        d += timedelta(days=1)
    return out


def compute_window(min_date: date, max_date: date) -> Tuple[date, date]:
    today = datetime.now().date()

    if DATE_PRESET == "THIS_WEEK":
        s = today - timedelta(days=today.weekday())
        e = s + timedelta(days=6)
        return s, e

    if DATE_PRESET == "ROLLING":
        s = today - timedelta(days=PAST_DAYS)
        e = today + timedelta(days=FUTURE_DAYS)
        return s, e

    if DATE_PRESET == "CUSTOM":
        s_in = parse_ymd(START_DATE)
        e_in = parse_ymd(END_DATE)

        if s_in and e_in:
            s, e = s_in, e_in
        elif s_in and not e_in:
            s, e = s_in, min(max_date, s_in + timedelta(days=MAX_DAYS - 1))
        elif (not s_in) and e_in:
            e = e_in
            s = max(min_date, e_in - timedelta(days=MAX_DAYS - 1))
        else:
            s = min_date
            e = min(max_date, s + timedelta(days=MAX_DAYS - 1))

        if (e - s).days + 1 > MAX_DAYS:
            e = s + timedelta(days=MAX_DAYS - 1)
        return s, e

    raise ValueError("DATE_PRESET 只支持 THIS_WEEK / ROLLING / CUSTOM")


def clip_interval(start_d: date, end_d: date, win_s: date, win_e: date) -> Tuple[Optional[date], Optional[date]]:
    s = max(start_d, win_s)
    e = min(end_d, win_e)
    if e < s:
        return None, None
    return s, e


def build_machine_list(tasks_win: List[dict]) -> List[str]:
    machines_in_data = sorted({t["machine"] for t in tasks_win})

    if MACHINE_ORDER:
        base = [m for m in MACHINE_ORDER if m not in HIDE_MACHINES]
        if INCLUDE_OTHER_MACHINES:
            base += [m for m in machines_in_data if m not in base and m not in HIDE_MACHINES]
        machines = base
    else:
        machines = [m for m in machines_in_data if m not in HIDE_MACHINES]

    if HIDE_EMPTY_MACHINES:
        has_task = {t["machine"] for t in tasks_win}
        machines = [m for m in machines if m in has_task]

    return machines


def assign_lanes(tasks_for_machine: List[dict], win_s: date, win_e: date) -> List[List[dict]]:
    """
    含结束日：不重叠条件 new_start > last_end
    """
    tasks_sorted = sorted(tasks_for_machine, key=lambda x: (x["start_ms"], x["end_ms"], x["task_id"]))
    lanes: List[List[dict]] = []
    lane_last_end: List[date] = []

    for t in tasks_sorted:
        s = parse_ymd(t["start_date"])
        e = parse_ymd(t["end_date"])
        s_eff, e_eff = clip_interval(s, e, win_s, win_e)
        if s_eff is None:
            continue

        placed = False
        for i in range(len(lanes)):
            if s_eff > lane_last_end[i]:
                lanes[i].append(t)
                lane_last_end[i] = e_eff
                placed = True
                break
        if not placed:
            lanes.append([t])
            lane_last_end.append(e_eff)

    return lanes or [[]]


def build_auto_col_width_map(tasks_win: List[dict], win_s: date, win_e: date) -> Dict[str, float]:
    """
    date.isoformat() -> column_width

    新规则：
    1) EXCLUDE_WIDTH_MACHINES 中的 machine 永远不参与列宽统计（即使单日也不算）
    2) 跨天任务（窗口内有效 span_days_inclusive > 1）完全不参与列宽统计
    3) 单日任务参与：该天 req_chars 取当日参与任务中 max_len 的最大值
    """
    dates = build_dates(win_s, win_e)
    req_chars: Dict[str, int] = {d.isoformat(): 0 for d in dates}

    excluded = {str(x).strip() for x in (EXCLUDE_WIDTH_MACHINES or []) if str(x).strip()}

    for t in tasks_win:
        m = str(t.get("machine", "") or "").strip()
        if m in excluded:
            continue

        s = parse_ymd(t["start_date"])
        e = parse_ymd(t["end_date"])
        s_eff, e_eff = clip_interval(s, e, win_s, win_e)
        if s_eff is None:
            continue

        span = (e_eff - s_eff).days + 1
        if span > 1:
            continue  # 跨天任务不参与

        max_len = int(t.get("max_len", 0) or 0)
        k = s_eff.isoformat()
        req_chars[k] = max(req_chars.get(k, 0), max_len)

    out: Dict[str, float] = {}
    for k, n in req_chars.items():
        w = (n + COL_PADDING_CHARS) * COL_CHAR_FACTOR
        w = max(DATE_COL_WIDTH_MIN, min(DATE_COL_WIDTH_MAX, w))
        out[k] = float(w)
    return out


def build_frame_xlsx(
    dates: List[date],
    machines: List[str],
    machine_lanes: Dict[str, List[List[dict]]],
    col_width_map: Optional[Dict[str, float]],
    out_xlsx: str,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "排产"
    ws.sheet_view.showGridLines = bool(SHOW_EXCEL_GRIDLINES)

    # 列宽
    ws.column_dimensions["A"].width = MACHINE_COL_WIDTH
    for i, d in enumerate(dates):
        col = get_column_letter(2 + i)
        if AUTO_DATE_COL_WIDTH and col_width_map:
            ws.column_dimensions[col].width = col_width_map.get(d.isoformat(), DATE_COL_WIDTH_MIN)
        else:
            ws.column_dimensions[col].width = DATE_COL_WIDTH_MIN

    # 边框Side（对齐参考文件）
    v_side = Side(style=DATE_VLINE_STYLE, color=DATE_VLINE_COLOR)
    a_sep_side = Side(style=MACHINE_DATE_VLINE_STYLE, color=MACHINE_DATE_VLINE_COLOR)
    sep_side = Side(style=MACHINE_SEP_STYLE, color=MACHINE_SEP_COLOR)

    # =========================
    # 样式模板（参考文件：排产视图_示例_无图例_v2.xlsx）
    # =========================
    # A列浅灰
    FILL_A = PatternFill("solid", fgColor="00F8FAFC")
    # 空白底板灰（rgb(243,244,246)）
    FILL_GRID = PatternFill("solid", fgColor="00F3F4F6")
    # 表头白
    FILL_WHITE = PatternFill("solid", fgColor="00FFFFFF")

    # 今天列：表头浅蓝 + 表体浅蓝
    FILL_TODAY_HEAD = PatternFill("solid", fgColor="00EFF6FF")
    FILL_TODAY_BODY = PatternFill("solid", fgColor="00E8F0FF")

    # 字体（你要求严格按参考的：A列、前两行日期）
    FONT_A_HEAD = Font(name="Microsoft YaHei", size=14, bold=True, color="00111827")
    FONT_A_MACHINE = Font(name="Microsoft YaHei", size=12, bold=True, color="00111827")

    FONT_DATE = Font(name="Microsoft YaHei", size=16, bold=True, color="00111827")
    FONT_WEEKDAY = Font(name="Microsoft YaHei", size=11, bold=False, color="00111827")

    FONT_DATE_TODAY = Font(name="Microsoft YaHei", size=16, bold=True, color="002563EB")
    FONT_WEEKDAY_TODAY = Font(name="Microsoft YaHei", size=11, bold=False, color="002563EB")
    TODAY_UNDERLINE = Side(style="medium", color="002563EB")

    # 计算今天列（不在窗口则 None）
    today = datetime.now().date()
    today_idx = next((i for i, d in enumerate(dates) if d == today), None)
    today_col = (2 + today_idx) if today_idx is not None else None

    # 标题（不改其他）
    ws["A1"] = TITLE
    ws["A1"].font = Font(bold=True, size=18)

    # 表头行高（不改其他）
    ws.row_dimensions[2].height = HEADER_ROW_HEIGHT
    ws.row_dimensions[3].height = HEADER_ROW_HEIGHT

    # A2:A3 机台表头
    ws["A2"] = "机台"
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].font = FONT_A_HEAD
    ws["A2"].fill = FILL_A
    ws.merge_cells("A2:A3")
    ws["A3"].fill = FILL_A

    # 日期表头（第2行：月/日；第3行：周几）
    for idx, d in enumerate(dates):
        c = 2 + idx
        is_today = (today_col == c)

        ws.cell(2, c, f"{d.month}/{d.day}").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(3, c, weekday_cn(d)).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(2, c).font = FONT_DATE_TODAY if is_today else FONT_DATE
        ws.cell(3, c).font = FONT_WEEKDAY_TODAY if is_today else FONT_WEEKDAY

        ws.cell(2, c).fill = FILL_TODAY_HEAD if is_today else FILL_WHITE
        ws.cell(3, c).fill = FILL_TODAY_HEAD if is_today else FILL_WHITE

        # 竖线：对齐参考文件（thin + 00E5E7EB）
        ws.cell(2, c).border = Border(left=v_side if c == 2 else None, right=v_side)
        ws.cell(3, c).border = Border(left=v_side if c == 2 else None, right=v_side)

        # 今天列：星期行蓝色下划线（参考文件 D2 的效果）
        if is_today:
            b = ws.cell(3, c).border
            ws.cell(3, c).border = Border(left=b.left, right=b.right, top=b.top, bottom=TODAY_UNDERLINE)

    # A列与日期区分割线（medium + 00E5E7EB）
    ws["A2"].border = Border(right=a_sep_side)

    # 冻结（不改其他）
    ws.freeze_panes = "B4"

    row = 4
    last_date_col = 1 + len(dates)

    # 机台块
    for m in machines:
        lanes = machine_lanes[m] or [[]]
        lane_count = len(lanes)
        start_row = row
        end_row = row + lane_count - 1

        # 机台名（合并）
        ws.cell(start_row, 1, m).font = FONT_A_MACHINE
        ws.cell(start_row, 1).alignment = Alignment(horizontal="center", vertical="center")
        if lane_count > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

        # 行高
        for r in range(start_row, end_row + 1):
            ws.row_dimensions[r].height = BASE_LANE_ROW_HEIGHT

        # 块边框：只保留竖线 + 机台块上下粗线（lane 内不画横线）
        for r in range(start_row, end_row + 1):
            top_side = sep_side if (APPLY_MACHINE_SEPARATORS and MACHINE_SEP_TOP and r == start_row) else None
            bottom_side = sep_side if (APPLY_MACHINE_SEPARATORS and MACHINE_SEP_BOTTOM and r == end_row) else None

            # A列：右分割线 + 块上下线（并套 A列底色）
            a_cell = ws.cell(r, 1)
            a_cell.border = Border(right=a_sep_side, top=top_side, bottom=bottom_side)
            a_cell.fill = FILL_A
            a_cell.alignment = Alignment(horizontal="center", vertical="center")

            # 日期列：空框架阶段 —— 全部铺底板灰；今天列整列覆盖浅蓝
            for c in range(2, last_date_col + 1):
                left_side = v_side if c == 2 else None
                right_side = v_side

                cell = ws.cell(r, c)
                cell.value = ""
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = Border(left=left_side, right=right_side, top=top_side, bottom=bottom_side)

                # ===== 背景规则（按你的三点要求修正）=====
                # 1) 排产区域底色：rgb(243,244,246) -> 00F3F4F6
                # 2) 今天列：整列覆盖（表体用 00E8F0FF）
                if today_col == c:
                    cell.fill = FILL_TODAY_BODY
                else:
                    cell.fill = FILL_GRID

        row = end_row + 1

    wb.save(out_xlsx)


def build_layout_json(
    dates: List[date],
    machines: List[str],
    machine_lanes: Dict[str, List[List[dict]]],
    win_s: date,
    win_e: date,
    out_layout: str,
):
    base_row = 4
    date_base_col = 2

    # 机台起始行
    machine_start_row: Dict[str, int] = {}
    r = base_row
    for m in machines:
        machine_start_row[m] = r
        r += len(machine_lanes[m] or [[]])

    items: List[dict] = []
    for m in machines:
        start_r = machine_start_row[m]
        for lane_idx, lane_tasks in enumerate(machine_lanes[m] or [[]]):
            row_num = start_r + lane_idx

            lane_tasks_sorted = sorted(lane_tasks, key=lambda x: (x["start_ms"], x["end_ms"], x["task_id"]))
            for t in lane_tasks_sorted:
                s = parse_ymd(t["start_date"])
                e = parse_ymd(t["end_date"])
                s_eff, e_eff = clip_interval(s, e, win_s, win_e)
                if s_eff is None:
                    continue

                start_c = date_base_col + (s_eff - win_s).days
                end_c = date_base_col + (e_eff - win_s).days

                items.append(
                    {
                        "task_id": t["task_id"],
                        "machine": m,
                        "lane": lane_idx,
                        "row": row_num,
                        "col_start": start_c,
                        "col_end": end_c,
                        "start_date": s_eff.isoformat(),
                        "end_date": e_eff.isoformat(),
                        "start_ms": t["start_ms"],
                        "end_ms": t["end_ms"],
                        "text_raw": t.get("text_raw", ""),
                        "line1": t.get("line1", ""),
                        "line2": t.get("line2", ""),
                        "line3": t.get("line3", ""),
                        "max_len": t.get("max_len", 0),
                        "rgb": t.get("rgb", None),
                    }
                )

    payload = {
        "title": TITLE,
        "win_start": win_s.isoformat(),
        "win_end": win_e.isoformat(),
        "dates": [d.isoformat() for d in dates],
        "machines": machines,
        "items": items,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    with open(out_layout, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def main():
    prepared_path = PREPARED_JSON
    if AUTO_LATEST:
        lp = pick_latest_prepared(OUTPUT_DIR)
        if lp:
            prepared_path = lp

    if not os.path.exists(prepared_path):
        raise FileNotFoundError(f"prepared 文件不存在：{prepared_path}")

    with open(prepared_path, "r", encoding="utf-8") as f:
        prepared = json.load(f)

    tasks = prepared.get("tasks", [])
    if not isinstance(tasks, list) or not tasks:
        raise RuntimeError("prepared tasks 为空：请先运行 B05A_prepare_data.py")

    # min/max 日期
    min_d = parse_ymd(prepared.get("min_date")) or min(parse_ymd(t["start_date"]) for t in tasks)
    max_d = parse_ymd(prepared.get("max_date")) or max(parse_ymd(t["end_date"]) for t in tasks)

    win_s, win_e = compute_window(min_d, max_d)
    dates = build_dates(win_s, win_e)

    # 窗口过滤
    tasks_win = []
    for t in tasks:
        s = parse_ymd(t["start_date"])
        e = parse_ymd(t["end_date"])
        if not (e < win_s or s > win_e):
            tasks_win.append(t)

    if not tasks_win:
        raise RuntimeError("窗口内无任务：请调整 DATE_PRESET/START_DATE/END_DATE")

    machines = build_machine_list(tasks_win)
    if not machines:
        raise RuntimeError("窗口内无机台：请检查隐藏/排序配置")

    # 按机台聚合
    by_machine: Dict[str, List[dict]] = {}
    for t in tasks_win:
        by_machine.setdefault(t["machine"], []).append(t)

    # lanes
    machine_lanes: Dict[str, List[List[dict]]] = {}
    for m in machines:
        machine_lanes[m] = assign_lanes(by_machine.get(m, []), win_s, win_e)

    # 自动列宽
    col_width_map = build_auto_col_width_map(tasks_win, win_s, win_e) if AUTO_DATE_COL_WIDTH else None

    # 输出
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_xlsx = os.path.join(OUTPUT_DIR, f"排产_框架_{ts}.xlsx")
    out_layout = os.path.join(OUTPUT_DIR, f"排产_layout_{ts}.json")

    build_frame_xlsx(dates, machines, machine_lanes, col_width_map, out_xlsx)
    build_layout_json(dates, machines, machine_lanes, win_s, win_e, out_layout)
    out_html = os.path.join(OUTPUT_DIR, f"schedule_web_{ts}.html")
    with open(out_layout, "r", encoding="utf-8") as f:
        layout_payload = json.load(f)
    layout_payload["today"] = datetime.now().date().isoformat()
    render_schedule_html(layout_payload, out_html)

    # 打印预览（少量）
    print(f"OK: {out_xlsx}")
    print(f"layout: {out_layout}")
    print(f"web: {out_html}")
    print(f"win: {win_s} ~ {win_e} | days={len(dates)} | machines={len(machines)} | tasks_in_win={len(tasks_win)}")
    head = sorted(tasks_win, key=lambda x: (x["machine"], x["start_ms"], x["end_ms"]))[:PRINT_FIRST_N]
    for i, t in enumerate(head, 1):
        rgb = "" if t.get("rgb") is None else ",".join(map(str, t["rgb"]))
        print(
            f"{i:02d} | {t['machine']} | {t['start_date']}~{t['end_date']} | "
            f"max_len={t.get('max_len',0)} | RGB={rgb or '空'} | "
            f"{t.get('line1','')}/{t.get('line2','')}/{t.get('line3','')}"
        )


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)
