# -*- coding: utf-8 -*-
import sys
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ================================
# 📂 配置
# ================================
DESKTOP = Path(r"C:\Users\ishel\Desktop\坚果备份\A四川和裕达新材料有限公司\24品质技术开发\运算文件夹")

MATRIX_FILE_GLOB = "物料清单合并_*.xlsx"
MATRIX_SHEET_NAME = "父子件比例矩阵"

SALES_FILE_GLOB = "销售价格波动分析表*.xlsx"
SALES_SRC_SHEET = "第一页"
SALES_DST_SHEET = "销售价格"

HEADER_ROW = 4              # 第4行是表头（矩阵表与销售表）
COL_NAME_CODE   = "存货编码"
COL_NAME_PRICE  = "最近价格1"
COL_NAME_NAME   = "存货名称"
COL_NAME_CUSTOM = "客户"

NUM_FMT_PRICE  = "0.00"
NUM_FMT_RATE   = "0.0000%"

# 期望在矩阵第4行出现的核心表头
EXPECTED_HEADERS = ["版本", "编码", "名称", "利润率", "利润", "售价", "成本"]

# ================================
# 🔧 日志 & 步骤包装
# ================================
def log(msg: str):
    print(msg if len(msg) <= 300 else (msg[:297] + "..."))

def step(title, fn, *args, **kwargs):
    log(f"▶ 开始：{title}")
    try:
        result, info = fn(*args, **kwargs)
        if isinstance(info, dict):
            brief = "; ".join(f"{k}={v}" for k, v in info.items())
        else:
            brief = str(info)
        log(f"✅ 成功：{title} ｜ {brief}")
        return result
    except Exception as e:
        log(f"❌ 失败：{title} ｜ {type(e).__name__}: {e}")
        raise

# ================================
# 🔧 基础工具
# ================================
def s(v):
    return "" if v is None else str(v).strip()

def to_float(x):
    try:
        v = pd.to_numeric(x)
        return float(v) if pd.notna(v) else None
    except Exception:
        return None

def find_latest(glob_pat: str):
    files = sorted(DESKTOP.glob(glob_pat), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"未找到文件：{glob_pat}")
    path = files[0]
    top3 = "; ".join(f.name for f in files[:3])
    return path, {"latest": path.name, "candidates(top3)": top3}

# ================================
# 🔧 矩阵表相关
# ================================
def open_matrix():
    path, _ = find_latest(MATRIX_FILE_GLOB)
    wb = load_workbook(path)
    if MATRIX_SHEET_NAME not in wb.sheetnames:
        raise KeyError(f"缺少工作表：{MATRIX_SHEET_NAME}")
    ws = wb[MATRIX_SHEET_NAME]
    return (wb, ws, path), {"file": path.name, "sheet": MATRIX_SHEET_NAME}

def map_headers(ws, row_idx: int):
    """读取指定行的表头，返回 {表头:列号}，并校验需要的字段是否存在。"""
    header = {}
    for c in range(1, ws.max_column + 1):
        key = s(ws.cell(row=row_idx, column=c).value)
        if key:
            header[key] = c
    missing = [h for h in EXPECTED_HEADERS if h not in header]
    if missing:
        raise KeyError(f"第{row_idx}行缺少表头: {missing}（应含{EXPECTED_HEADERS}）")
    info = {k: get_column_letter(header[k]) for k in ["编码", "售价", "成本", "利润率", "利润"]}
    info["cols"] = len(header)
    return header, info

def data_region(ws, header):
    """根据矩阵表头计算数据区起点（信息性）。"""
    header_cols_max = max(header[h] for h in EXPECTED_HEADERS)
    start_col = header_cols_max + 1
    for c in range(start_col, ws.max_column + 1):
        if s(ws.cell(row=1, column=c).value):
            start_col = c
            break
    start_row = HEADER_ROW + 1  # 通常 = 5
    return (start_row, start_col), {"start_cell": f"{get_column_letter(start_col)}{start_row}"}

# ================================
# 🔧 销售价格表相关
# ================================
def copy_sales_sheet(wb):
    sales_path, _ = find_latest(SALES_FILE_GLOB)
    src_wb = load_workbook(sales_path, data_only=True)
    if SALES_SRC_SHEET not in src_wb.sheetnames:
        raise KeyError(f"销售簿缺少工作表：{SALES_SRC_SHEET}")
    if SALES_DST_SHEET in wb.sheetnames:
        del wb[SALES_DST_SHEET]
    dst = wb.create_sheet(title=SALES_DST_SHEET)
    src = src_wb[SALES_SRC_SHEET]
    rows = 0
    for r in src.iter_rows(values_only=True):
        dst.append(list(r))
        rows += 1
    return SALES_DST_SHEET, {"copied_rows": rows, "from_file": sales_path.name}

def parse_sales_headers(ws_sales):
    """解析销售表第4行的表头，定位 存货编码 / 存货名称 / 客户 / 最近价格1 列。"""
    header = {}
    for c in range(1, ws_sales.max_column + 1):
        key = s(ws_sales.cell(row=HEADER_ROW, column=c).value)
        if key:
            header[key] = c
    col_code  = header.get(COL_NAME_CODE)
    col_price = header.get(COL_NAME_PRICE)
    col_name  = header.get(COL_NAME_NAME)      # 允许缺失
    col_cust  = header.get(COL_NAME_CUSTOM)    # 允许缺失
    if not col_code or not col_price:
        raise KeyError(f"销售表缺少列：‘{COL_NAME_CODE}’ 或 ‘{COL_NAME_PRICE}’（表头行={HEADER_ROW}）")
    info = {
        "code_col":  get_column_letter(col_code),
        "price_col": get_column_letter(col_price),
        "name_col":  get_column_letter(col_name) if col_name else "-",
        "cust_col":  get_column_letter(col_cust) if col_cust else "-",
    }
    return (col_code, col_price, col_name, col_cust), info

def build_price_map_max(ws_sales, col_code, col_price, col_name=None, col_cust=None):
    """
    构建 “存货编码 → 最近价格1(取最大)” 映射。
    如果同一编码出现多个不同价格：
      - 逐条打印候选（存货编码 / 存货名称 / 客户 / 最近价格1）
      - 打印最终选择的最大值（并标注对应名称/客户）
    """
    # 收集：code -> list of dict(price, name, cust)
    buckets = {}
    for r in range(HEADER_ROW + 1, ws_sales.max_row + 1):
        code = s(ws_sales.cell(row=r, column=col_code).value)
        price = to_float(ws_sales.cell(row=r, column=col_price).value)
        if not code or price is None:
            continue
        name = s(ws_sales.cell(row=r, column=col_name).value) if col_name else "—"
        cust = s(ws_sales.cell(row=r, column=col_cust).value) if col_cust else "—"
        buckets.setdefault(code, []).append({"price": price, "name": name, "cust": cust})

    chosen = {}
    multi = 0
    for code, items in buckets.items():
        if not items:
            continue
        # 最大价格
        mx = max(x["price"] for x in items)
        # 触发多价打印（不同价格）
        if len(set(x["price"] for x in items)) > 1:
            multi += 1
            log(f"ℹ️ 多价选择触发：存货编码={code}")
            # 候选逐条打印
            for it in items:
                log(f"   候选｜存货编码={code}｜存货名称={it['name']}｜客户={it['cust']}｜最近价格1={it['price']}")
            # 标注最终选择（优先第一条匹配最大值的记录）
            best = next(it for it in items if it["price"] == mx)
            log(f"   → 取最大｜存货编码={code}｜存货名称={best['name']}｜客户={best['cust']}｜最近价格1={best['price']}")
        chosen[code] = mx

    return chosen, {"codes": len(chosen), "duplicates": multi}

def write_sales_to_matrix(ws_matrix, matrix_hdr, price_map):
    """把销售价格写入矩阵：按“编码”列值匹配，写入“售价”列。"""
    col_encode = matrix_hdr["编码"]
    col_sale   = matrix_hdr["售价"]
    filled = 0
    for r in range(HEADER_ROW + 1, ws_matrix.max_row + 1):
        code = s(ws_matrix.cell(row=r, column=col_encode).value)
        if code and code in price_map:
            cell = ws_matrix.cell(row=r, column=col_sale, value=price_map[code])
            cell.number_format = NUM_FMT_PRICE
            filled += 1
    return filled, {"written_rows": filled, "col_code": get_column_letter(col_encode), "col_sale": get_column_letter(col_sale)}

def calc_profit_rate_and_profit(ws_matrix, matrix_hdr):
    """
    重新计算：
      利润率 = (售价 - 成本) / 成本   （成本<=0或缺失：跳过）
      利润   = 售价 - 成本           （售价或成本缺失：跳过）
    """
    col_sale   = matrix_hdr["售价"]
    col_cost   = matrix_hdr["成本"]
    col_rate   = matrix_hdr["利润率"]
    col_profit = matrix_hdr["利润"]

    filled_rate = 0
    filled_profit = 0
    skipped_rate = 0
    skipped_profit = 0

    for r in range(HEADER_ROW + 1, ws_matrix.max_row + 1):
        sale = to_float(ws_matrix.cell(row=r, column=col_sale).value)
        cost = to_float(ws_matrix.cell(row=r, column=col_cost).value)

        # 利润
        if sale is not None and cost is not None:
            p = sale - cost
            ws_matrix.cell(row=r, column=col_profit, value=p).number_format = NUM_FMT_PRICE
            filled_profit += 1
        else:
            skipped_profit += 1

        # 利润率
        if sale is not None and cost is not None and cost != 0:
            rate = (sale - cost) / cost
            ws_matrix.cell(row=r, column=col_rate, value=rate).number_format = NUM_FMT_RATE
            filled_rate += 1
        else:
            skipped_rate += 1

    return (filled_rate, filled_profit), {
        "rate_ok": filled_rate, "rate_skip": skipped_rate,
        "profit_ok": filled_profit, "profit_skip": skipped_profit,
        "cols": f"rate={get_column_letter(col_rate)}, profit={get_column_letter(col_profit)}"
    }

def save_wb(wb, path: Path):
    try:
        wb.save(path)
        return None, {"saved": str(path)}
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = path.with_name(path.stem + f"_sales_{ts}" + path.suffix)
        wb.save(alt)
        return None, {"saved_alt": str(alt), "reason": "original file locked"}

# ================================
# 🚀 主流程
# ================================
def main():
    # 1) 打开矩阵文件与表
    (wb, ws, matrix_path) = step("定位并打开矩阵文件", open_matrix)

    # 2) 读取矩阵第4行表头并校验
    matrix_hdr = step("读取矩阵表头(第4行)并校验字段", map_headers, ws, HEADER_ROW)

    # 3) 计算矩阵数据区起点（信息性）
    _ = step("定位矩阵数据区起点", data_region, ws, matrix_hdr)

    # 4) 复制销售价格表为《销售价格》
    copied_name = step("复制销售表《第一页》为《销售价格》", copy_sales_sheet, wb)

    # 5) 解析销售表表头（第4行）
    ws_sales = wb[copied_name]
    (col_code, col_price, col_name, col_cust) = step("解析销售表表头(第4行)", parse_sales_headers, ws_sales)

    # 6) 构建“存货编码→最近价格1(最大)”映射，并打印多价选择过程（带名称/客户）
    price_map = step("构建售价映射（多价取最大并打印名称/客户/价格）",
                     build_price_map_max, ws_sales, col_code, col_price, col_name, col_cust)

    # 7) 写入矩阵“售价”列（按“编码”列匹配）
    _ = step("写入矩阵‘售价’列（按‘编码’匹配）", write_sales_to_matrix, ws, matrix_hdr, price_map)

    # 8) 重新计算“利润率/利润”
    _ = step("计算‘利润率’与‘利润’列", calc_profit_rate_and_profit, ws, matrix_hdr)

    # 9) 保存
    _ = step("保存工作簿", save_wb, wb, matrix_path)

    log("🏁 全流程完成。")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"❌ 终止：{type(e).__name__}: {e}")
        sys.exit(1)
