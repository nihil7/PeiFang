# insert_cols_rows.py
import sys
import time
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ================================
# 📂 配置路径 / 工作表名（集中）
# ================================
DESKTOP = Path(r"C:\Users\ishel\Desktop\坚果备份\A四川和裕达新材料有限公司\24品质技术开发\运算文件夹")
INPUT_GLOB = "物料清单合并_*.xlsx"
SHEET_MATRIX = "父子件比例矩阵"
HISTORY_GLOB = "上一次物料编码及名称*.xlsx"
HISTORY_SHEET = "上一次物料编码及名称"

HEADERS_ROW4 = ["版本", "编码", "名称", "利润率", "利润", "售价", "成本"]

# ================================
# 🔊 日志工具（统一风格 + 用时）
# ================================
def _pack_info(info):
    if not info:
        return "-"
    if isinstance(info, dict):
        parts = []
        for k, v in info.items():
            parts.append(f"{k}={v}")
        return "; ".join(parts)
    return str(info)

def log_start(title):
    print(f"▶ 开始：{title}")

def log_ok(title, info=None, elapsed=None):
    t = f" | {elapsed:.3f}s" if elapsed is not None else ""
    print(f"✅ 成功：{title}{t} ｜ {_pack_info(info)}")

def log_warn(msg):
    print(f"⚠️ 提示：{msg}")

def log_fail(title, err, elapsed=None):
    t = f" | {elapsed:.3f}s" if elapsed is not None else ""
    print(f"❌ 失败：{title}{t} ｜ {type(err).__name__}: {err}")

def step(title, fn, *args, **kwargs):
    log_start(title)
    t0 = time.perf_counter()
    try:
        result, info = fn(*args, **kwargs)
        log_ok(title, info, time.perf_counter() - t0)
        return result
    except Exception as e:
        log_fail(title, e, time.perf_counter() - t0)
        raise

# ================================
# 🔧 工具函数
# ================================
def s(v):
    """安全转字符串+strip：None→""，其他类型先转str再strip"""
    return "" if v is None else str(v).strip()

def find_latest(glob_pat: str):
    files = sorted(DESKTOP.glob(glob_pat), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"未找到文件：{glob_pat}")
    path = files[0]
    info = {"latest": path.name}
    # 打印前3个候选，便于核对
    if len(files) > 1:
        info["candidates(top3)"] = "; ".join(f.name for f in files[:3])
    return path, info

def load_matrix_wb(path: Path, sheet_name: str):
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"缺少工作表《{sheet_name}》，请先运行程序一。")
    ws = wb[sheet_name]
    return (wb, ws), {"sheets": ", ".join(wb.sheetnames), "target": sheet_name}

def already_processed(ws):
    vals = [ s(ws.cell(row=4, column=i).value) for i in range(1, 8) ]
    ok = (vals == HEADERS_ROW4)
    info = {"A4:G4": "/".join(vals), "matched": ok}
    return ok, info

def insert_rows_cols_and_style(ws):
    rows_before, cols_before = ws.max_row, ws.max_column

    # 插入：第2行后插2行；第3列后插4列（openpyxl在 idx 之前插入）
    ws.insert_rows(idx=3, amount=2)
    ws.insert_cols(idx=4, amount=4)

    # 第4行表头（A4:G4）
    for col, text in enumerate(HEADERS_ROW4, start=1):
        cell = ws.cell(row=4, column=col, value=text)
        cell.alignment = Alignment(horizontal="center")

    # 数据区左上角：原 D3 → 插入后为 H5
    data_r0, data_c0 = 5, 8  # H=8

    # 冻结窗格到 H5
    ws.freeze_panes = ws["H5"]

    # 列宽与对齐
    widths = [16, 18, 22, 10, 10, 10, 10]  # A..G
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 数据区列宽（H列及以后）
    for col in range(8, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10

    # 顶部两行（子件编码/名称）居中
    for c in range(1, ws.max_column + 1):
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=c).alignment = Alignment(horizontal="center")

    info = {
        "rows(before→after)": f"{rows_before}→{ws.max_row}",
        "cols(before→after)": f"{cols_before}→{ws.max_column}",
        "frozen": "H5",
        "header_row4": "/".join(HEADERS_ROW4),
        "data_origin": f"{get_column_letter(data_c0)}{data_r0}"
    }
    return (data_r0, data_c0), info

def copy_history_sheet(wb):
    try:
        hist_path, info_latest = find_latest(HISTORY_GLOB)
        hwb = load_workbook(hist_path, data_only=True)
        if HISTORY_SHEET not in hwb.sheetnames:
            log_warn("历史簿缺少指定工作表，跳过复制")
            return None, {"copied": False, "reason": "missing sheet"}
        # 若重名，自动追加(1)(2)...
        name = HISTORY_SHEET
        idx = 1
        while name in wb.sheetnames:
            idx += 1
            name = f"{HISTORY_SHEET}({idx})"
        dst_ws = wb.create_sheet(title=name)
        src_ws = hwb[HISTORY_SHEET]
        rows = 0
        for r in src_ws.iter_rows(values_only=True):
            dst_ws.append(list(r)); rows += 1
        return name, {"copied": True, "from": hist_path.name, "to": name, "rows": rows}
    except Exception as e:
        # 可选步骤：失败不中断主流程
        log_warn(f"未复制历史表：{type(e).__name__}: {e}")
        return None, {"copied": False, "error": str(e)}

def fill_std_name(ws, history_ws_name, wb):
    """
    将历史表中的“标准名称型号”写到第4行、对应子件列（H列及以后）。
    - 子件编码位于第1行（build_matrix 产出）
    """
    if not history_ws_name:
        return 0, {"filled": 0, "reason": "no history sheet"}

    hws = wb[history_ws_name]

    # 第1行“子件编码”→列号映射（统一转字符串）
    code2col = {}
    for c in range(1, ws.max_column + 1):
        key = s(ws.cell(row=1, column=c).value)
        if key:
            code2col[key] = c

    # 历史表头定位（第1行，统一转字符串）
    header = { s(hws.cell(row=1, column=c).value): c
               for c in range(1, hws.max_column + 1) }
    col_code = header.get("子件编码")
    col_std  = header.get("标准名称型号") or header.get("标准名称型号 ")
    if not col_code or not col_std:
        return 0, {"filled": 0, "reason": "missing columns(子件编码/标准名称型号)"}

    filled = 0
    tried = 0
    for r in range(2, hws.max_row + 1):
        code = s(hws.cell(row=r, column=col_code).value)
        stdn = hws.cell(row=r, column=col_std).value
        if not code:
            continue
        tried += 1
        if code in code2col:
            col_idx = code2col[code]
            # 仅在子件列（H列及以后）写入，避免覆盖 A4..G4
            if col_idx >= 8:
                ws.cell(row=4, column=col_idx, value=stdn)
                filled += 1
    info = {"tried": tried, "filled": filled, "mapped_cols": len(code2col)}
    return filled, info

# === 新增：按历史“标准名称型号”顺序重排子件列（仅对已匹配到名称的列生效） ===
def reorder_cols_by_history_std(ws, history_ws_name, wb):
    """
    依据历史表的“标准名称型号”列顺序，对矩阵中子件区(H列及以后)那些在第4行已写入
    “标准名称型号”的列进行重排；未匹配到的列保持相对顺序不变。
    发现重复或不匹配时打印统计信息与示例。
    """
    if not history_ws_name:
        return False, {"reordered": False, "reason": "no history sheet"}

    hws = wb[history_ws_name]
    # 找到“标准名称型号”列
    header = { s(hws.cell(row=1, column=c).value): c
               for c in range(1, hws.max_column + 1) }
    col_std = header.get("标准名称型号") or header.get("标准名称型号 ")
    if not col_std:
        return False, {"reordered": False, "reason": "history missing 标准名称型号"}

    # 历史顺序列表
    hist_order = []
    seen_hist = set()
    hist_dupes = set()
    for r in range(2, hws.max_row + 1):
        name = s(hws.cell(row=r, column=col_std).value)
        if not name:
            continue
        if name in seen_hist:
            hist_dupes.add(name)
            # 仍保留第一次出现的顺序，不追加
        else:
            seen_hist.add(name)
            hist_order.append(name)

    # 当前子件区列信息（从H列=8开始）
    first_col = 8
    max_col = ws.max_column
    max_row = ws.max_row

    # 收集：每列的标准名称型号（取第4行）
    col_records = []
    name2cols = {}
    for c in range(first_col, max_col + 1):
        std_name = s(ws.cell(row=4, column=c).value)
        # 采集整列数据（值 + number_format，以便重写后尽量保持单元格格式）
        col_vals = []
        col_nf = []
        for r in range(1, max_row + 1):
            cell = ws.cell(row=r, column=c)
            col_vals.append(cell.value)
            col_nf.append(cell.number_format)
        rec = {"col": c, "name": std_name, "vals": col_vals, "nfs": col_nf}
        col_records.append(rec)
        if std_name:
            name2cols.setdefault(std_name, []).append(rec)

    # 重复/未匹配统计
    ws_dupes = {n: recs for n, recs in name2cols.items() if len(recs) > 1}
    # 目标顺序：先放“历史中出现且当前表已匹配”的列，按照历史顺序；若某名称对应多列，则保留这些列的原先相对顺序
    matched_recs = []
    missing_in_ws = []
    for name in hist_order:
        recs = name2cols.get(name)
        if recs:
            matched_recs.extend(recs)
        else:
            missing_in_ws.append(name)

    # 未匹配（当前ws里有，但历史里没有或第4行为空）的列，保持原相对顺序追加在后
    matched_set = set(id(r) for r in matched_recs)
    unmatched_recs = [r for r in col_records if id(r) not in matched_set]

    new_order = matched_recs + unmatched_recs

    # 回写列（覆盖 H..last 的内容与 number_format）
    for i, rec in enumerate(new_order):
        tgt_c = first_col + i
        for r in range(1, max_row + 1):
            cell = ws.cell(row=r, column=tgt_c)
            cell.value = rec["vals"][r-1]
            # 尽量保留格式（第8步会重刷百分比）
            cell.number_format = rec["nfs"][r-1]

    # 统计与简短示例（避免日志过长）
    def _sample(names, n=5):
        return ", ".join(names[:n]) + ("..." if len(names) > n else "")

    info = {
        "matched_cols": len(matched_recs),
        "unmatched_cols": len(unmatched_recs),
        "ws_dupes_cnt": len(ws_dupes),
        "hist_dupes_cnt": len(hist_dupes),
    }
    if ws_dupes:
        info["ws_dupes(sample)"] = _sample(list(ws_dupes.keys()))
    if hist_dupes:
        info["hist_dupes(sample)"] = _sample(sorted(hist_dupes))
    if missing_in_ws:
        info["missing_in_ws(sample)"] = _sample(missing_in_ws)

    return True, info

def reapply_percent_format(ws, data_r0, data_c0):
    # 对 H5 起的数据区再次统一百分比格式
    count = 0
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(data_r0, max_row + 1):
        for c in range(data_c0, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)):
                ws.cell(row=r, column=c).number_format = "0.0000%"
                count += 1
    return count, {"cells_formatted": count, "range_from": f"{get_column_letter(data_c0)}{data_r0}"}

def set_autofilter_row4(ws):
    # 将筛选范围覆盖整表，从 A4 到 最后一列/最后一行
    ref = f"A4:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.auto_filter.ref = ref
    return ref, {"auto_filter": ref}

def set_g123_headers(ws):
    """
    不插/删任何单元格：G1=原料编码, G2=原料名称, G3=原料含税单价
    """
    before = [ws["G1"].value, ws["G2"].value, ws["G3"].value]
    ws["G1"].value = "原料编码"
    ws["G2"].value = "原料名称"
    ws["G3"].value = "原料含税单价"
    ws["G1"].alignment = Alignment(horizontal="center")
    ws["G2"].alignment = Alignment(horizontal="center")
    ws["G3"].alignment = Alignment(horizontal="center")
    after = [ws["G1"].value, ws["G2"].value, ws["G3"].value]
    return True, {"G1~G3(before→after)": f"{before}→{after}"}

def save_wb(wb, path: Path):
    try:
        wb.save(path)
        return path.name, {"saved": path.name}
    except PermissionError as e:
        # 被占用则另存
        alt = path.with_name(path.stem + "_structlog" + path.suffix)
        wb.save(alt)
        return alt.name, {"saved_alt": alt.name, "reason": "original file locked"}

# ================================
# 🚀 主流程
# ================================
def main():
    # 1) 锁定输入文件
    input_path = step("查找最新输入文件", find_latest, INPUT_GLOB)

    # 2) 打开工作簿/工作表
    (wb, ws) = step("加载工作簿并定位矩阵表", load_matrix_wb, input_path, SHEET_MATRIX)

    # 3) 是否已处理过（A4:G4）
    processed = step("检测是否已存在目标表头(A4:G4)", already_processed, ws)
    if processed:
        # 3.1 已处理：仅补齐 G1~G3 & 筛选，并保存
        step("写入G1~G3表头（不插不删）", set_g123_headers, ws)
        step("设置筛选范围(第4行表头)", set_autofilter_row4, ws)
        step("保存工作簿", save_wb, wb, input_path)
        print("🏁 完成：检测到已处理过，本次执行仅同步G1~G3与筛选。")
        return

    # 4) 首次处理：插入行列+表头+样式
    (data_r0, data_c0) = step("插入2行/4列并设置样式与冻结", insert_rows_cols_and_style, ws)

    # 5) 写 G1~G3（原料编码/名称/含税单价）
    step("写入G1~G3表头（不插不删）", set_g123_headers, ws)

    # 6) 复制历史表（可选步骤，失败不中断）
    history_name = step("复制历史表到当前簿（可选）", copy_history_sheet, wb)

    # 7) 回填“标准名称型号”到第4行相应子件列
    step("回填历史‘标准名称型号’到第4行（子件列）", fill_std_name, ws, history_name, wb)

    # 7.5) **新增**：按历史“标准名称型号”顺序重排子件列（仅对已匹配列）
    step("按历史‘标准名称型号’顺序重排子件列", reorder_cols_by_history_std, ws, history_name, wb)

    # 8) 统一百分比格式（H5起）
    step("统一数据区百分比格式（H5起）", reapply_percent_format, ws, data_r0, data_c0)

    # 9) 设置筛选（以第4行表头覆盖整表）
    step("设置筛选范围(第4行表头)", set_autofilter_row4, ws)

    # 10) 保存
    step("保存工作簿", save_wb, wb, input_path)

    print("🏁 完成：首次处理全部步骤执行完毕。")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"🛑 异常终止：{type(e).__name__}: {e}")
        sys.exit(1)
