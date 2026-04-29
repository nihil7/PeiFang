"""
程序简介：保留历史流程或实验逻辑，仅供追溯参考，主流程优先使用 apps 或 tools 下的新入口。
主要逻辑：读取所需配置或输入数据，执行本文件负责的处理步骤，并把结果写入本地文件或输出到命令行。
配置说明：涉及企微或飞书凭证时，优先读取 PEIFANG_ENV_PROFILE、WECOM_ENV_PROFILE、FEISHU_ENV_PROFILE 选择公司配置档案；未设置时兼容原来的 .env 变量。
"""

# B04.py
# 作用：Bitable JSON -> Excel 排产（日历式）
# 重点：任务用“圆角矩形 Shape”（可编辑文字、可随单元格缩放），允许覆盖网格线
# 依赖：pip install openpyxl pywin32 pandas

import os
import re
import sys
import json
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

# ============= 配置区（只改这里） =============
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

AUTO_LATEST = True
FIELDS_JSON = os.path.join(OUTPUT_DIR, "生产任务排期__排产·统计总台账__20260114_230447.fields.json")
RECORDS_JSON = os.path.join(OUTPUT_DIR, "生产任务排期__排产·统计总台账__20260114_230447.records.raw.json")

TITLE = "机台×日期排程（日历式）"

# 日期范围：THIS_WEEK / ROLLING / CUSTOM
DATE_PRESET = "THIS_WEEK"
PAST_DAYS = 3
FUTURE_DAYS = 14
START_DATE = None  # "2026-01-05" 或 None
END_DATE = None    # "2026-01-20" 或 None
MAX_DAYS = 31

# 机台排序/显示
MACHINE_ORDER: List[str] = ["35机", "4#65机", "1#机", "2#机"]
INCLUDE_OTHER_MACHINES = False
HIDE_EMPTY_MACHINES = True
HIDE_MACHINES: List[str] = []

# 表格视觉
MACHINE_COL_WIDTH = 14
DATE_COL_WIDTH = 18
HEADER_ROW_HEIGHT = 26
BASE_LANE_ROW_HEIGHT = 44  # 每条lane基础行高（可手动调大/调小）

# Shape（圆角卡片）视觉
SHAPE_MARGIN = 4          # 允许覆盖网格线：留小边距更好看
SHAPE_LINE_COLOR = 0xA0A0A0
SHAPE_LINE_WEIGHT = 1.5
SHAPE_RADIUS = 0.22       # 圆角程度（0~1，越大越圆）
SHAPE_FONT_SIZE = 12
SHAPE_BOLD = True

# 打印核对
PRINT_FIRST_N = 20
# ============================================


def ensure_dir(p: str):
    os.makedirs(p, exist_ok=True)


def pick_latest_pair(output_dir: str) -> Tuple[str, str]:
    files = os.listdir(output_dir)
    fields = [f for f in files if f.endswith(".fields.json")]
    if not fields:
        raise FileNotFoundError(f"在 {output_dir} 找不到 *.fields.json")
    fields.sort(key=lambda fn: os.path.getmtime(os.path.join(output_dir, fn)), reverse=True)
    latest_fields = fields[0]
    prefix = latest_fields[:-len(".fields.json")]
    latest_records = prefix + ".records.raw.json"
    fields_path = os.path.join(output_dir, latest_fields)
    records_path = os.path.join(output_dir, latest_records)

    if not os.path.exists(records_path):
        records = [f for f in files if f.endswith(".records.raw.json")]
        if not records:
            raise FileNotFoundError(f"在 {output_dir} 找不到 *.records.raw.json")
        records.sort(key=lambda fn: os.path.getmtime(os.path.join(output_dir, fn)), reverse=True)
        records_path = os.path.join(output_dir, records[0])

    return fields_path, records_path


def first_text_cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, list) and v:
        if isinstance(v[0], dict) and "text" in v[0]:
            return str(v[0].get("text", "")).strip()
    if isinstance(v, str):
        return v.strip()
    return ""


def clamp_rgb(r: int, g: int, b: int) -> Tuple[int, int, int]:
    return max(0, min(255, r)), max(0, min(255, g)), max(0, min(255, b))


def parse_rgb(s: Any) -> Optional[Tuple[int, int, int]]:
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    m = re.match(r"rgb\(\s*(\d{1,3})\s*,\s*(\d{1,3})\s*,\s*(\d{1,3})\s*\)", s, re.I)
    if m:
        return clamp_rgb(*map(int, m.groups()))
    m = re.match(r"#?([0-9a-fA-F]{6})$", s)
    if m:
        h = m.group(1)
        return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    m = re.match(r"(\d{1,3})\s*,\s*(\d{1,3})\s*,\s*(\d{1,3})$", s)
    if m:
        return clamp_rgb(*map(int, m.groups()))
    return None


def ms_to_date(ms: Any):
    if ms is None or ms == "":
        return None
    ms_int = int(ms)
    dt = datetime.fromtimestamp(ms_int / 1000, tz=timezone.utc)
    return dt.date()


def weekday_cn(d):
    names = ["一", "二", "三", "四", "五", "六", "日"]
    return f"周{names[d.weekday()]}"


def parse_ymd(x: Optional[str]):
    if not x:
        return None
    return datetime.strptime(x, "%Y-%m-%d").date()


def compute_window(tasks_all: List[dict]):
    today = datetime.now().date()
    data_min = min(t["start"] for t in tasks_all)
    data_max = max(t["end"] for t in tasks_all)

    if DATE_PRESET == "THIS_WEEK":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return start, end

    if DATE_PRESET == "ROLLING":
        start = today - timedelta(days=PAST_DAYS)
        end = today + timedelta(days=FUTURE_DAYS)
        return start, end

    if DATE_PRESET == "CUSTOM":
        s_in = parse_ymd(START_DATE)
        e_in = parse_ymd(END_DATE)

        if s_in and e_in:
            start, end = s_in, e_in
            if (end - start).days + 1 > MAX_DAYS:
                end = start + timedelta(days=MAX_DAYS - 1)
            return start, end

        if s_in and (e_in is None):
            start = s_in
            end = min(data_max, start + timedelta(days=MAX_DAYS - 1))
            return start, end

        if (s_in is None) and e_in:
            end = e_in
            start = max(data_min, end - timedelta(days=MAX_DAYS - 1))
            return start, end

        start = data_min
        end = min(data_max, start + timedelta(days=MAX_DAYS - 1))
        return start, end

    raise ValueError("DATE_PRESET 只支持 THIS_WEEK / ROLLING / CUSTOM")


def build_dates(start, end) -> List[Any]:
    out = []
    d = start
    while d <= end:
        out.append(d)
        d += timedelta(days=1)
    return out


def extract_tasks(records_doc: dict) -> List[dict]:
    tasks = []
    for rec in records_doc.get("records", []):
        vals = rec.get("values", {}) or {}

        # 机台
        machine = ""
        mv = vals.get("机台")
        if isinstance(mv, list) and mv and isinstance(mv[0], dict):
            machine = str(mv[0].get("text", "")).strip()

        text = first_text_cell(vals.get("甘特图文本"))
        rgb_tuple = parse_rgb(first_text_cell(vals.get("RGB")))

        s_ms = vals.get("开始日期")
        e_ms = vals.get("结束日期")
        if s_ms in (None, "") or e_ms in (None, ""):
            continue
        try:
            s_ms = int(s_ms)
            e_ms = int(e_ms)
        except Exception:
            continue

        s = ms_to_date(s_ms)
        e = ms_to_date(e_ms)
        if not machine or not text or not s or not e:
            continue

        if e_ms < s_ms:
            s_ms, e_ms = e_ms, s_ms
            s, e = e, s

        rid = rec.get("record_id", "")
        task_id = f"{rid}_{s_ms}_{e_ms}" if rid else f"{machine}_{s_ms}_{e_ms}_{abs(hash(text))}"

        tasks.append({
            "task_id": task_id,
            "machine": machine,
            "start": s,
            "end": e,
            "start_ms": s_ms,
            "end_ms": e_ms,
            "text": text,
            "rgb": rgb_tuple,  # None => 透明，仅边框
        })
    return tasks


def clip_interval(t: dict, win_start, win_end):
    s = max(t["start"], win_start)
    e = min(t["end"], win_end)
    if e < s:
        return None, None
    return s, e


def assign_lanes(tasks_for_machine: List[dict], win_start, win_end) -> List[List[dict]]:
    # 结束日含当天：不重叠要求 new_start > last_end
    tasks_sorted = sorted(tasks_for_machine, key=lambda x: (x["start_ms"], x["end_ms"], x["text"]))
    lanes: List[List[dict]] = []
    lane_last_end: List[Any] = []

    for t in tasks_sorted:
        s_eff, e_eff = clip_interval(t, win_start, win_end)
        if s_eff is None:
            continue

        placed = False
        for i in range(len(lanes)):
            if s_eff > lane_last_end[i]:
                lanes[i].append(t)
                lane_last_end[i] = e_eff
                placed = True
                break
        if not placed:
            lanes.append([t])
            lane_last_end.append(e_eff)

    return lanes


def build_machine_list(tasks_in_window: List[dict]) -> List[str]:
    machines_in_data = sorted({t["machine"] for t in tasks_in_window})

    if MACHINE_ORDER:
        base = [m for m in MACHINE_ORDER if m not in HIDE_MACHINES]
        if INCLUDE_OTHER_MACHINES:
            base += [m for m in machines_in_data if m not in base and m not in HIDE_MACHINES]
        machines = base
    else:
        machines = [m for m in machines_in_data if m not in HIDE_MACHINES]

    if HIDE_EMPTY_MACHINES:
        has_task = {t["machine"] for t in tasks_in_window}
        machines = [m for m in machines if m in has_task]

    return machines


def rgb_to_ole(rgb: Tuple[int, int, int]) -> int:
    # Excel OLE_COLOR = BGR
    r, g, b = rgb
    return (b << 16) | (g << 8) | r


def print_preview(tasks_in_win: List[dict], win_start, win_end):
    print(f"窗口：{win_start} ~ {win_end}（含结束日）")
    print(f"窗口内任务数：{len(tasks_in_win)}")
    head = sorted(tasks_in_win, key=lambda x: (x["machine"], x["start_ms"], x["end_ms"]))[:PRINT_FIRST_N]
    for i, t in enumerate(head, 1):
        rgb = "" if t["rgb"] is None else f"{t['rgb'][0]},{t['rgb'][1]},{t['rgb'][2]}"
        span = (t["end"] - t["start"]).days + 1
        print(f"{i:02d} | {t['machine']} | {t['start']}~{t['end']}({span}天) | RGB={rgb or '空'} | {t['text'][:40]}")


def write_check_csv(tasks_in_win: List[dict], out_csv: str):
    rows = []
    for t in tasks_in_win:
        rows.append({
            "机台": t["machine"],
            "开始(ms)": t["start_ms"],
            "结束(ms)": t["end_ms"],
            "开始(日期)": t["start"].isoformat(),
            "结束(日期)": t["end"].isoformat(),
            "跨度天数(含结束日)": (t["end"] - t["start"]).days + 1,
            "RGB": "" if t["rgb"] is None else f"{t['rgb'][0]},{t['rgb'][1]},{t['rgb'][2]}",
            "甘特图文本": t["text"],
        })
    pd.DataFrame(rows).to_csv(out_csv, index=False, encoding="utf-8-sig")


def build_grid_xlsx(dates: List[Any], machines: List[str], machine_lanes: Dict[str, List[List[dict]]], out_xlsx: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "排产"

    # 列宽
    ws.column_dimensions["A"].width = MACHINE_COL_WIDTH
    for i in range(len(dates)):
        ws.column_dimensions[get_column_letter(2 + i)].width = DATE_COL_WIDTH

    # 边框
    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 标题
    ws["A1"] = TITLE
    ws["A1"].font = Font(bold=True, size=18)

    # 表头（两行）
    ws.row_dimensions[2].height = HEADER_ROW_HEIGHT
    ws.row_dimensions[3].height = HEADER_ROW_HEIGHT

    ws["A2"] = "机台"
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].font = Font(bold=True, size=12)
    ws.merge_cells("A2:A3")

    for idx, d in enumerate(dates):
        c = 2 + idx
        ws.cell(2, c, f"{d.month}/{d.day}").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(3, c, weekday_cn(d)).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(2, c).font = Font(bold=True, size=12)
        ws.cell(3, c).font = Font(bold=True, size=12)

    # 网格与机台+lanes
    row = 4
    for m in machines:
        lanes = machine_lanes[m]
        lane_count = len(lanes)

        start_row = row
        end_row = row + lane_count - 1

        # 机台名（合并）
        ws.cell(start_row, 1, m).font = Font(bold=True, size=12)
        ws.cell(start_row, 1).alignment = Alignment(horizontal="left", vertical="center")
        if lane_count > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

        # 画网格单元
        for r in range(start_row, end_row + 1):
            ws.row_dimensions[r].height = BASE_LANE_ROW_HEIGHT
            for c in range(1, 2 + len(dates)):
                cell = ws.cell(r, c)
                cell.border = border
                if c != 1:  # 日期区域
                    cell.value = ""
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        row = end_row + 1

    # 冻结窗格（B4）
    ws.freeze_panes = "B4"

    wb.save(out_xlsx)


def add_shapes_with_excel(dates: List[Any], machines: List[str], machine_lanes: Dict[str, List[List[dict]]],
                          win_start, win_end, xlsx_path: str):
    try:
        import win32com.client  # type: ignore
    except Exception:
        raise RuntimeError("缺少 pywin32：请先 pip install pywin32")

    # Excel 常量
    msoShapeRoundedRectangle = 5
    msoTrue = -1
    xlMoveAndSize = 1

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
    ws = wb.Worksheets("排产")

    # 从第4行开始，按 build_grid_xlsx 的逻辑一致
    base_row = 4
    date_base_col = 2  # B列

    # 机台起始行映射
    machine_start_row: Dict[str, int] = {}
    r = base_row
    for m in machines:
        machine_start_row[m] = r
        r += len(machine_lanes[m])

    # 插形状
    for m in machines:
        start_r = machine_start_row[m]
        lanes = machine_lanes[m]

        for lane_idx, lane_tasks in enumerate(lanes):
            row_num = start_r + lane_idx

            # 任务排序：早的在上
            lane_tasks_sorted = sorted(lane_tasks, key=lambda x: (x["start_ms"], x["end_ms"], x["text"]))

            for t in lane_tasks_sorted:
                s_eff, e_eff = clip_interval(t, win_start, win_end)
                if s_eff is None:
                    continue

                start_c = date_base_col + (s_eff - win_start).days
                end_c = date_base_col + (e_eff - win_start).days

                # 用 Range 获取跨列区域的像素尺寸
                rng = ws.Range(ws.Cells(row_num, start_c), ws.Cells(row_num, end_c))
                left = rng.Left + SHAPE_MARGIN
                top = rng.Top + SHAPE_MARGIN
                width = max(10, rng.Width - 2 * SHAPE_MARGIN)
                height = max(10, rng.Height - 2 * SHAPE_MARGIN)

                shp = ws.Shapes.AddShape(msoShapeRoundedRectangle, left, top, width, height)
                shp.Placement = xlMoveAndSize
                shp.Adjustments[1] = SHAPE_RADIUS


                # 边框
                shp.Line.Visible = msoTrue
                shp.Line.ForeColor.RGB = SHAPE_LINE_COLOR
                shp.Line.Weight = SHAPE_LINE_WEIGHT

                # 填充
                if t["rgb"] is None:
                    shp.Fill.Visible = msoTrue
                    shp.Fill.Transparency = 1.0  # 透明
                else:
                    shp.Fill.Visible = msoTrue
                    shp.Fill.Transparency = 0.0
                    shp.Fill.ForeColor.RGB = rgb_to_ole(t["rgb"])

                # 文本（可编辑）
                tf = shp.TextFrame2
                tf.WordWrap = msoTrue
                tf.MarginLeft = 6
                tf.MarginRight = 6
                tf.MarginTop = 4
                tf.MarginBottom = 4

                tr = tf.TextRange
                tr.Text = t["text"]
                tr.Font.Size = SHAPE_FONT_SIZE
                tr.Font.Bold = SHAPE_BOLD
                tr.ParagraphFormat.Alignment = 1  # 左对齐
                tr.ParagraphFormat.FirstLineIndent = 0

    wb.Save()
    wb.Close(SaveChanges=True)
    excel.Quit()


def main():
    ensure_dir(OUTPUT_DIR)

    fields_path = FIELDS_JSON
    records_path = RECORDS_JSON
    if AUTO_LATEST:
        fields_path, records_path = pick_latest_pair(OUTPUT_DIR)

    if not os.path.exists(records_path):
        raise FileNotFoundError(records_path)

    with open(records_path, "r", encoding="utf-8") as f:
        records_doc = json.load(f)

    tasks_all = extract_tasks(records_doc)
    if not tasks_all:
        raise RuntimeError("未抽取到任务：请确认 records.raw.json 中包含 机台/开始日期/结束日期/甘特图文本")

    win_start, win_end = compute_window(tasks_all)
    dates = build_dates(win_start, win_end)

    tasks_in_win = [t for t in tasks_all if not (t["end"] < win_start or t["start"] > win_end)]
    machines = build_machine_list(tasks_in_win)
    if not machines:
        raise RuntimeError("窗口内无机台任务：请调整日期范围")

    # 按机台分组
    tasks_by_machine: Dict[str, List[dict]] = {}
    for t in tasks_in_win:
        tasks_by_machine.setdefault(t["machine"], []).append(t)

    # lanes
    machine_lanes: Dict[str, List[List[dict]]] = {}
    for m in machines:
        machine_lanes[m] = assign_lanes(tasks_by_machine.get(m, []), win_start, win_end) or [[]]

    # 打印核对（关键字段）
    print_preview(tasks_in_win, win_start, win_end)

    # 输出核对CSV
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    check_csv = os.path.join(OUTPUT_DIR, f"生产任务排期_核对_{ts}.csv")
    write_check_csv(tasks_in_win, check_csv)

    # 先生成网格xlsx
    out_xlsx = os.path.join(OUTPUT_DIR, f"生产任务排期_圆角Shape_{ts}.xlsx")
    build_grid_xlsx(dates, machines, machine_lanes, out_xlsx)

    # 再用Excel插入圆角Shape
    add_shapes_with_excel(dates, machines, machine_lanes, win_start, win_end, out_xlsx)

    print(f"OK: {out_xlsx}")
    print(f"核对CSV: {check_csv}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)
