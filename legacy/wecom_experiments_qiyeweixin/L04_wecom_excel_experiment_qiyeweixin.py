"""
程序简介：保留历史流程或实验逻辑，仅供追溯参考，主流程优先使用 apps 或 tools 下的新入口。
主要逻辑：读取所需配置或输入数据，执行本文件负责的处理步骤，并把结果写入本地文件或输出到命令行。
配置说明：涉及企微或飞书凭证时，优先读取 PEIFANG_ENV_PROFILE、WECOM_ENV_PROFILE、FEISHU_ENV_PROFILE 选择公司配置档案；未设置时兼容原来的 .env 变量。
"""

# B03_制作排产Excel.py
# 作用：从 Bitable 导出的 fields/records JSON 生成“日期为列、机台为行(带lanes)、跨天合并、RGB填色”的 Excel 排产表
# 规则：
# - 结束日含当天（跨天=合并单元格）
# - 同机台同一天多任务：自动分配到不同 lanes（子行）
# - 机台名用合并单元格（类似 rowSpan）
# - RGB 为空：不填充，只保留边框；RGB 有值：填充颜色
# - 自动换行，并按列宽估算行高撑开（近似 Excel 自动适配效果）

import os
import re
import sys
import json
import math
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter


# =========================
# 配置区（只改这里）
# =========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

AUTO_LATEST = True
FIELDS_JSON = os.path.join(OUTPUT_DIR, "生产任务排期__排产·统计总台账__20260114_230447.fields.json")
RECORDS_JSON = os.path.join(OUTPUT_DIR, "生产任务排期__排产·统计总台账__20260114_230447.records.raw.json")

TITLE = "机台×日期排程（Excel）"

# 日期范围：THIS_WEEK / ROLLING / CUSTOM
DATE_PRESET = "THIS_WEEK"
PAST_DAYS = 3
FUTURE_DAYS = 14
START_DATE = None  # "2026-01-05" 或 None
END_DATE = None    # "2026-01-20" 或 None
MAX_DAYS = 31

# 机台排序/显示
MACHINE_ORDER: List[str] = ["35机", "4#65机", "1#机", "2#机"]
INCLUDE_OTHER_MACHINES = False
HIDE_EMPTY_MACHINES = True
HIDE_MACHINES: List[str] = []

# 表格样式
MACHINE_COL_WIDTH = 14
DATE_COL_WIDTH = 18
BASE_ROW_HEIGHT = 22  # 基础行高（1行文字）
PADDING_LINES = 0.3   # 行高额外留白（行数系数）

# =========================
# 基础工具
# =========================
def ensure_dir(p: str):
    os.makedirs(p, exist_ok=True)

def pick_latest_pair(output_dir: str) -> Tuple[str, str]:
    files = os.listdir(output_dir)
    fields = [f for f in files if f.endswith(".fields.json")]
    if not fields:
        raise FileNotFoundError(f"在 {output_dir} 找不到 *.fields.json")
    fields.sort(key=lambda fn: os.path.getmtime(os.path.join(output_dir, fn)), reverse=True)
    latest_fields = fields[0]
    prefix = latest_fields[:-len(".fields.json")]
    latest_records = prefix + ".records.raw.json"
    latest_fields_path = os.path.join(output_dir, latest_fields)
    latest_records_path = os.path.join(output_dir, latest_records)

    if not os.path.exists(latest_records_path):
        records = [f for f in files if f.endswith(".records.raw.json")]
        if not records:
            raise FileNotFoundError(f"在 {output_dir} 找不到 *.records.raw.json")
        records.sort(key=lambda fn: os.path.getmtime(os.path.join(output_dir, fn)), reverse=True)
        latest_records_path = os.path.join(output_dir, records[0])

    return latest_fields_path, latest_records_path

def first_text_cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, list) and v:
        if isinstance(v[0], dict) and "text" in v[0]:
            return str(v[0].get("text", "")).strip()
    if isinstance(v, str):
        return v.strip()
    return ""

def clamp_rgb(r: int, g: int, b: int) -> Tuple[int, int, int]:
    return max(0, min(255, r)), max(0, min(255, g)), max(0, min(255, b))

def parse_rgb(s: Any) -> Optional[Tuple[int, int, int]]:
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    m = re.match(r"rgb\(\s*(\d{1,3})\s*,\s*(\d{1,3})\s*,\s*(\d{1,3})\s*\)", s, re.I)
    if m:
        return clamp_rgb(*map(int, m.groups()))
    m = re.match(r"#?([0-9a-fA-F]{6})$", s)
    if m:
        h = m.group(1)
        return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    m = re.match(r"(\d{1,3})\s*,\s*(\d{1,3})\s*,\s*(\d{1,3})$", s)
    if m:
        return clamp_rgb(*map(int, m.groups()))
    return None

def ms_to_date(ms: Any):
    if ms is None or ms == "":
        return None
    try:
        ms_int = int(ms)
    except Exception:
        return None
    dt = datetime.fromtimestamp(ms_int / 1000, tz=timezone.utc)
    return dt.date()

def weekday_cn(d):
    names = ["一", "二", "三", "四", "五", "六", "日"]
    return f"周{names[d.weekday()]}"

def parse_ymd(x: Optional[str]):
    if not x:
        return None
    return datetime.strptime(x, "%Y-%m-%d").date()

def compute_window(tasks_all: List[dict]):
    today = datetime.now().date()
    data_min = min(t["start"] for t in tasks_all)
    data_max = max(t["end"] for t in tasks_all)

    if DATE_PRESET == "THIS_WEEK":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return start, end

    if DATE_PRESET == "ROLLING":
        start = today - timedelta(days=PAST_DAYS)
        end = today + timedelta(days=FUTURE_DAYS)
        return start, end

    if DATE_PRESET == "CUSTOM":
        s_in = parse_ymd(START_DATE)
        e_in = parse_ymd(END_DATE)

        if s_in and e_in:
            start, end = s_in, e_in
            if (end - start).days + 1 > MAX_DAYS:
                end = start + timedelta(days=MAX_DAYS - 1)
            return start, end

        if s_in and (e_in is None):
            start = s_in
            end = min(data_max, start + timedelta(days=MAX_DAYS - 1))
            return start, end

        if (s_in is None) and e_in:
            end = e_in
            start = max(data_min, end - timedelta(days=MAX_DAYS - 1))
            return start, end

        start = data_min
        end = min(data_max, start + timedelta(days=MAX_DAYS - 1))
        return start, end

    raise ValueError("DATE_PRESET 只支持 THIS_WEEK / ROLLING / CUSTOM")

def build_dates(start, end) -> List[Any]:
    out = []
    d = start
    while d <= end:
        out.append(d)
        d += timedelta(days=1)
    return out


# =========================
# 数据抽取 & lanes
# =========================
def extract_tasks(records_doc: dict) -> List[dict]:
    tasks = []
    for rec in records_doc.get("records", []):
        vals = rec.get("values", {}) or {}

        machine = ""
        mv = vals.get("机台")
        if isinstance(mv, list) and mv and isinstance(mv[0], dict):
            machine = str(mv[0].get("text", "")).strip()

        text = first_text_cell(vals.get("甘特图文本"))
        rgb_tuple = parse_rgb(first_text_cell(vals.get("RGB")))

        s_ms = vals.get("开始日期")
        e_ms = vals.get("结束日期")
        if s_ms in (None, "") or e_ms in (None, ""):
            continue
        try:
            s_ms = int(s_ms)
            e_ms = int(e_ms)
        except Exception:
            continue

        s = ms_to_date(s_ms)
        e = ms_to_date(e_ms)
        if not machine or not text or not s or not e:
            continue

        if e_ms < s_ms:
            s_ms, e_ms = e_ms, s_ms
            s, e = e, s

        rid = rec.get("record_id", "")
        task_id = f"{rid}_{s_ms}_{e_ms}" if rid else f"{machine}_{s_ms}_{e_ms}_{abs(hash(text))}"

        tasks.append({
            "task_id": task_id,
            "machine": machine,
            "start": s,
            "end": e,
            "start_ms": s_ms,
            "end_ms": e_ms,
            "text": text,
            "rgb": rgb_tuple,  # None => 不填充
        })
    return tasks

def clip_interval(t: dict, win_start, win_end):
    s = max(t["start"], win_start)
    e = min(t["end"], win_end)
    if e < s:
        return None, None
    return s, e

def assign_lanes(tasks_for_machine: List[dict], win_start, win_end) -> List[List[dict]]:
    # 同机台排序：早的在上
    tasks_sorted = sorted(tasks_for_machine, key=lambda x: (x["start_ms"], x["end_ms"], x["text"]))
    lanes: List[List[dict]] = []
    lane_last_end: List[Any] = []

    for t in tasks_sorted:
        s_eff, e_eff = clip_interval(t, win_start, win_end)
        if s_eff is None:
            continue

        placed = False
        for i in range(len(lanes)):
            # 结束日含当天：不重叠要求 new_start > last_end
            if s_eff > lane_last_end[i]:
                lanes[i].append(t)
                lane_last_end[i] = e_eff
                placed = True
                break
        if not placed:
            lanes.append([t])
            lane_last_end.append(e_eff)

    return lanes

def build_machine_list(tasks_in_window: List[dict]) -> List[str]:
    machines_in_data = sorted({t["machine"] for t in tasks_in_window})

    if MACHINE_ORDER:
        base = [m for m in MACHINE_ORDER if m not in HIDE_MACHINES]
        if INCLUDE_OTHER_MACHINES:
            base += [m for m in machines_in_data if m not in base and m not in HIDE_MACHINES]
        machines = base
    else:
        machines = [m for m in machines_in_data if m not in HIDE_MACHINES]

    if HIDE_EMPTY_MACHINES:
        has_task = {t["machine"] for t in tasks_in_window}
        machines = [m for m in machines if m in has_task]

    return machines


# =========================
# Excel 绘制
# =========================
def rgb_to_fill(rgb: Tuple[int, int, int]) -> PatternFill:
    # openpyxl 需要 ARGB
    r, g, b = rgb
    argb = f"FF{r:02X}{g:02X}{b:02X}"
    return PatternFill(fill_type="solid", start_color=argb, end_color=argb)

def set_border_range(ws, r1, c1, r2, c2, border: Border):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = border

def estimate_lines(text: str, col_char_width: int) -> int:
    # 近似估算：按列宽字符数折行（中文/英文都粗略当1字符）
    if not text:
        return 1
    width = max(6, int(col_char_width))
    # 手动换行优先
    parts = str(text).splitlines() or [str(text)]
    lines = 0
    for p in parts:
        p = p.strip()
        if not p:
            lines += 1
        else:
            lines += math.ceil(len(p) / width)
    return max(1, lines)

def main():
    ensure_dir(OUTPUT_DIR)

    fields_path = FIELDS_JSON
    records_path = RECORDS_JSON
    if AUTO_LATEST:
        fields_path, records_path = pick_latest_pair(OUTPUT_DIR)

    if not os.path.exists(records_path):
        raise FileNotFoundError(records_path)

    with open(records_path, "r", encoding="utf-8") as f:
        records_doc = json.load(f)

    tasks_all = extract_tasks(records_doc)
    if not tasks_all:
        raise RuntimeError("未抽取到任务：请确认 records.raw.json 中包含 机台/开始日期/结束日期/甘特图文本")

    win_start, win_end = compute_window(tasks_all)
    dates = build_dates(win_start, win_end)

    # 与窗口有交集就保留
    tasks_in_win = [t for t in tasks_all if not (t["end"] < win_start or t["start"] > win_end)]
    machines = build_machine_list(tasks_in_win)
    if not machines:
        raise RuntimeError("窗口内无机台任务：请调整日期范围")

    # 机台 -> tasks
    tasks_by_machine: Dict[str, List[dict]] = {}
    for t in tasks_in_win:
        tasks_by_machine.setdefault(t["machine"], []).append(t)

    # lanes
    machine_lanes: Dict[str, List[List[dict]]] = {}
    for m in machines:
        machine_lanes[m] = assign_lanes(tasks_by_machine.get(m, []), win_start, win_end) or [[]]

    # 建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "排产"

    # 列宽
    ws.column_dimensions["A"].width = MACHINE_COL_WIDTH
    for i, d in enumerate(dates, start=2):  # B开始
        col = get_column_letter(i)
        ws.column_dimensions[col].width = DATE_COL_WIDTH

    # 样式
    thin = Side(style="thin", color="9E9E9E")
    thick = Side(style="medium", color="7A7A7A")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick_top = Border(left=thin, right=thin, top=thick, bottom=thin)
    border_thick_bottom = Border(left=thin, right=thin, top=thin, bottom=thick)
    border_thick_tb = Border(left=thin, right=thin, top=thick, bottom=thick)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    font_head = Font(bold=True, size=12)
    font_cell = Font(bold=True, size=12)


    # 表头两行（机台列）
    ws["A1"] = "机台"
    ws["A1"].alignment = align_center
    ws["A1"].font = font_head
    ws.merge_cells("A1:A2")  # ✅ 合并后不要再写 A2.value

    for idx, d in enumerate(dates):
        c = 2 + idx
        ws.cell(1, c, f"{d.month}/{d.day}")
        ws.cell(2, c, weekday_cn(d))
        ws.cell(1, c).alignment = align_center
        ws.cell(2, c).alignment = align_center
        ws.cell(1, c).font = font_head
        ws.cell(2, c).font = font_head

    # 画边框（先把表头边框铺一遍）
    max_col = 1 + len(dates)
    set_border_range(ws, 1, 1, 2, max_col, border_thin)

    # 内容行开始
    row = 3

    # 用于行高估算：按每行最大文字行数来撑高
    # （DATE_COL_WIDTH 大致可当每行可容纳字符数）
    chars_per_line = int(DATE_COL_WIDTH)

    for m in machines:
        lanes = machine_lanes[m]
        lane_count = len(lanes)

        start_row_machine = row
        end_row_machine = row + lane_count - 1

        # 机台名合并（rowSpan）
        ws.cell(start_row_machine, 1, m)
        ws.cell(start_row_machine, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.cell(start_row_machine, 1).font = font_head
        if lane_count > 1:
            ws.merge_cells(start_row=start_row_machine, start_column=1, end_row=end_row_machine, end_column=1)

        # 每条 lane 一行
        for lane_idx in range(lane_count):
            # 先铺空格与边框
            for ci in range(2, max_col + 1):
                ws.cell(row, ci, "")
                ws.cell(row, ci).alignment = align_left_wrap
                ws.cell(row, ci).font = font_cell
                ws.cell(row, ci).border = border_thin

            # 写任务（跨天 merge）
            # 任务按 start_ms 排序（保证更早的在上面）
            lane_tasks = sorted(lanes[lane_idx], key=lambda x: (x["start_ms"], x["end_ms"], x["text"]))

            # 该行的最大行数估计（决定行高）
            row_max_lines = 1

            for t in lane_tasks:
                s_eff, e_eff = clip_interval(t, win_start, win_end)
                if s_eff is None:
                    continue

                # 找到日期列索引（B=2）
                s_idx = (s_eff - win_start).days
                e_idx = (e_eff - win_start).days
                c1 = 2 + s_idx
                c2 = 2 + e_idx

                # 合并跨天
                if c2 > c1:
                    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)

                cell = ws.cell(row, c1)
                cell.value = t["text"]
                cell.alignment = align_left_wrap
                cell.font = font_cell

                # 填充
                if t["rgb"] is not None:
                    cell.fill = rgb_to_fill(t["rgb"])
                # RGB为空：不填充（保持默认）

                # 合并区域边框：给区域内所有格子边框统一（避免 merge 后丢边）
                for cc in range(c1, c2 + 1):
                    ws.cell(row, cc).border = border_thin

                # 行高估算：按合并后的“总宽度”估算可容纳字符
                span_cols = (c2 - c1 + 1)
                effective_width = max(6, chars_per_line * span_cols)
                row_max_lines = max(row_max_lines, estimate_lines(t["text"], effective_width))

            # 设置行高（自动换行撑高的近似实现）
            # Excel真正“自动适配”通常需要手动双击行边界；这里用估算先撑开
            ws.row_dimensions[row].height = BASE_ROW_HEIGHT * (row_max_lines + PADDING_LINES)

            row += 1

        # 机台块上下粗线：整块范围
        # 顶部粗线
        set_border_range(ws, start_row_machine, 1, start_row_machine, max_col, border_thick_top)
        # 底部粗线
        set_border_range(ws, end_row_machine, 1, end_row_machine, max_col, border_thick_bottom)

        # 机台列边框也铺一遍
        for rr in range(start_row_machine, end_row_machine + 1):
            ws.cell(rr, 1).border = border_thin
            ws.cell(rr, 1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.cell(rr, 1).font = font_head

    # 冻结窗格（方便滚动）
    ws.freeze_panes = "B3"

    # 保存核对 JSON（窗口内）
    check = []
    for t in tasks_in_win:
        check.append({
            "machine": t["machine"],
            "start_date": t["start"].isoformat(),
            "end_date": t["end"].isoformat(),
            "text": t["text"],
            "rgb": None if t["rgb"] is None else list(t["rgb"]),
        })
    with open(os.path.join(OUTPUT_DIR, "生产任务排期_Excel_核对.json"), "w", encoding="utf-8") as f:
        json.dump(check, f, ensure_ascii=False, indent=2)

    # 输出文件名
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_xlsx = os.path.join(OUTPUT_DIR, f"生产任务排期_Excel_排产_{ts}.xlsx")
    wb.save(out_xlsx)

    # 打印控制（简短）
    print(f"OK: {out_xlsx}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)

