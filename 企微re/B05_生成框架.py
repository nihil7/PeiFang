# B05_生成框架.py
# 作用：读取 output 下最新一组 records.raw.json -> 按北京时间解析日期 -> 计算 lanes -> 生成 Excel 框架 + layout.json + 核对CSV
# 依赖：pip install openpyxl pandas

import os
import re
import sys
import json
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

try:
    from zoneinfo import ZoneInfo
    BJ_TZ = ZoneInfo("Asia/Shanghai")
except Exception:
    BJ_TZ = timezone(timedelta(hours=8))


# =========================
# 配置区（只改这里）
# =========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

AUTO_LATEST = True
RECORDS_JSON = os.path.join(OUTPUT_DIR, "生产任务排期__排产·统计总台账__20260114_230447.records.raw.json")

TITLE = "机台×日期排程（框架）"

# 日期范围：THIS_WEEK / ROLLING / CUSTOM
DATE_PRESET = "ROLLING"
PAST_DAYS = 0
FUTURE_DAYS = 2
START_DATE = None  # "2026-01-05" 或 None
END_DATE = None    # "2026-01-20" 或 None
MAX_DAYS = 31

# 机台排序/显示
MACHINE_ORDER: List[str] = ["35机", "4#65机", "1#机", "2#机"]
INCLUDE_OTHER_MACHINES = False
HIDE_EMPTY_MACHINES = True
HIDE_MACHINES: List[str] = []

# 表格视觉
MACHINE_COL_WIDTH = 14
DATE_COL_WIDTH = 18
HEADER_ROW_HEIGHT = 26
BASE_LANE_ROW_HEIGHT = 44

# ——【新增】机台分隔线样式（框架里粗线就是这里控制）——
APPLY_MACHINE_SEPARATORS = True
MACHINE_SEP_STYLE = "medium"     # thin / medium / thick
MACHINE_SEP_COLOR = "8A8A8A"     # 16进制RGB（无#）
MACHINE_SEP_TOP = True          # 每个机台块的“上边框”加粗
MACHINE_SEP_BOTTOM = True       # 每个机台块的“下边框”加粗

# 普通网格线
GRID_STYLE = "thin"
GRID_COLOR = "B0B0B0"

# 输出控制
PRINT_FIRST_N = 20
# =========================


def ensure_dir(p: str):
    os.makedirs(p, exist_ok=True)


def pick_latest_records(output_dir: str) -> str:
    files = os.listdir(output_dir)
    recs = [f for f in files if f.endswith(".records.raw.json")]
    if not recs:
        raise FileNotFoundError(f"在 {output_dir} 找不到 *.records.raw.json")
    recs.sort(key=lambda fn: os.path.getmtime(os.path.join(output_dir, fn)), reverse=True)
    return os.path.join(output_dir, recs[0])


def first_text_cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, list) and v:
        if isinstance(v[0], dict) and "text" in v[0]:
            return str(v[0].get("text", "")).strip()
    if isinstance(v, str):
        return v.strip()
    return ""


def clamp_rgb(r: int, g: int, b: int) -> Tuple[int, int, int]:
    return max(0, min(255, r)), max(0, min(255, g)), max(0, min(255, b))


def parse_rgb(s: Any) -> Optional[Tuple[int, int, int]]:
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    m = re.match(r"rgb\(\s*(\d{1,3})\s*,\s*(\d{1,3})\s*,\s*(\d{1,3})\s*\)", s, re.I)
    if m:
        return clamp_rgb(*map(int, m.groups()))
    m = re.match(r"#?([0-9a-fA-F]{6})$", s)
    if m:
        h = m.group(1)
        return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    m = re.match(r"(\d{1,3})\s*,\s*(\d{1,3})\s*,\s*(\d{1,3})$", s)
    if m:
        return clamp_rgb(*map(int, m.groups()))
    return None


def ms_to_date_bj(ms: Any):
    if ms is None or ms == "":
        return None
    ms_int = int(ms)
    dt_utc = datetime.fromtimestamp(ms_int / 1000, tz=timezone.utc)
    dt_bj = dt_utc.astimezone(BJ_TZ)
    return dt_bj.date()


def weekday_cn(d):
    names = ["一", "二", "三", "四", "五", "六", "日"]
    return f"周{names[d.weekday()]}"


def parse_ymd(x: Optional[str]):
    if not x:
        return None
    return datetime.strptime(x, "%Y-%m-%d").date()


def extract_tasks(records_doc: dict) -> List[dict]:
    tasks = []
    for rec in records_doc.get("records", []):
        vals = rec.get("values", {}) or {}

        machine = ""
        mv = vals.get("机台")
        if isinstance(mv, list) and mv and isinstance(mv[0], dict):
            machine = str(mv[0].get("text", "")).strip()

        text = first_text_cell(vals.get("甘特图文本"))
        rgb_tuple = parse_rgb(first_text_cell(vals.get("RGB")))

        s_ms = vals.get("开始日期")
        e_ms = vals.get("结束日期")
        if s_ms in (None, "") or e_ms in (None, ""):
            continue
        try:
            s_ms = int(s_ms)
            e_ms = int(e_ms)
        except Exception:
            continue

        # ✅ 按北京时间取 date（修正你遇到的“差一天”）
        s = ms_to_date_bj(s_ms)
        e = ms_to_date_bj(e_ms)

        if not machine or not text or not s or not e:
            continue

        if e_ms < s_ms:
            s_ms, e_ms = e_ms, s_ms
            s, e = e, s

        rid = rec.get("record_id", "")
        task_id = f"{rid}_{s_ms}_{e_ms}" if rid else f"{machine}_{s_ms}_{e_ms}_{abs(hash(text))}"

        tasks.append({
            "task_id": task_id,
            "machine": machine,
            "start": s,
            "end": e,
            "start_ms": s_ms,
            "end_ms": e_ms,
            "text": text,
            "rgb": rgb_tuple,  # None 允许
        })
    return tasks


def compute_window(tasks_all: List[dict]):
    today = datetime.now(BJ_TZ).date()
    data_min = min(t["start"] for t in tasks_all)
    data_max = max(t["end"] for t in tasks_all)

    if DATE_PRESET == "THIS_WEEK":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return start, end

    if DATE_PRESET == "ROLLING":
        start = today - timedelta(days=PAST_DAYS)
        end = today + timedelta(days=FUTURE_DAYS)
        return start, end

    if DATE_PRESET == "CUSTOM":
        s_in = parse_ymd(START_DATE)
        e_in = parse_ymd(END_DATE)

        if s_in and e_in:
            start, end = s_in, e_in
            if (end - start).days + 1 > MAX_DAYS:
                end = start + timedelta(days=MAX_DAYS - 1)
            return start, end

        if s_in and (e_in is None):
            start = s_in
            end = min(data_max, start + timedelta(days=MAX_DAYS - 1))
            return start, end

        if (s_in is None) and e_in:
            end = e_in
            start = max(data_min, end - timedelta(days=MAX_DAYS - 1))
            return start, end

        start = data_min
        end = min(data_max, start + timedelta(days=MAX_DAYS - 1))
        return start, end

    raise ValueError("DATE_PRESET 只支持 THIS_WEEK / ROLLING / CUSTOM")


def build_dates(start, end) -> List[Any]:
    out = []
    d = start
    while d <= end:
        out.append(d)
        d += timedelta(days=1)
    return out


def clip_interval(t: dict, win_start, win_end):
    s = max(t["start"], win_start)
    e = min(t["end"], win_end)
    if e < s:
        return None, None
    return s, e


def assign_lanes(tasks_for_machine: List[dict], win_start, win_end) -> List[List[dict]]:
    # ✅ 含结束日：不重叠条件 new_start > last_end
    tasks_sorted = sorted(tasks_for_machine, key=lambda x: (x["start_ms"], x["end_ms"], x["text"]))
    lanes: List[List[dict]] = []
    lane_last_end: List[Any] = []

    for t in tasks_sorted:
        s_eff, e_eff = clip_interval(t, win_start, win_end)
        if s_eff is None:
            continue

        placed = False
        for i in range(len(lanes)):
            if s_eff > lane_last_end[i]:
                lanes[i].append(t)
                lane_last_end[i] = e_eff
                placed = True
                break
        if not placed:
            lanes.append([t])
            lane_last_end.append(e_eff)

    return lanes


def build_machine_list(tasks_in_window: List[dict]) -> List[str]:
    machines_in_data = sorted({t["machine"] for t in tasks_in_window})

    if MACHINE_ORDER:
        base = [m for m in MACHINE_ORDER if m not in HIDE_MACHINES]
        if INCLUDE_OTHER_MACHINES:
            base += [m for m in machines_in_data if m not in base and m not in HIDE_MACHINES]
        machines = base
    else:
        machines = [m for m in machines_in_data if m not in HIDE_MACHINES]

    if HIDE_EMPTY_MACHINES:
        has_task = {t["machine"] for t in tasks_in_window}
        machines = [m for m in machines if m in has_task]

    return machines


def print_preview(tasks_in_win: List[dict], win_start, win_end):
    print(f"窗口(北京)：{win_start}~{win_end} | 任务={len(tasks_in_win)}")
    head = sorted(tasks_in_win, key=lambda x: (x["machine"], x["start_ms"], x["end_ms"]))[:PRINT_FIRST_N]
    for i, t in enumerate(head, 1):
        rgb = "" if t["rgb"] is None else f"{t['rgb'][0]},{t['rgb'][1]},{t['rgb'][2]}"
        span = (t["end"] - t["start"]).days + 1
        print(f"{i:02d} | {t['machine']} | {t['start']}~{t['end']}({span}天) | RGB={rgb or '空'} | {t['text'][:40]}")


def write_check_csv(tasks_in_win: List[dict], out_csv: str):
    rows = []
    for t in tasks_in_win:
        rows.append({
            "机台": t["machine"],
            "开始(ms)": t["start_ms"],
            "结束(ms)": t["end_ms"],
            "开始(北京日期)": t["start"].isoformat(),
            "结束(北京日期)": t["end"].isoformat(),
            "跨度天数(含结束日)": (t["end"] - t["start"]).days + 1,
            "RGB": "" if t["rgb"] is None else f"{t['rgb'][0]},{t['rgb'][1]},{t['rgb'][2]}",
            "甘特图文本": t["text"],
        })
    pd.DataFrame(rows).to_csv(out_csv, index=False, encoding="utf-8-sig")


def apply_border_keep(cell, top=None, bottom=None, left=None, right=None):
    b = cell.border
    cell.border = Border(
        left=left if left is not None else b.left,
        right=right if right is not None else b.right,
        top=top if top is not None else b.top,
        bottom=bottom if bottom is not None else b.bottom,
    )


def build_frame_xlsx(dates: List[Any], machines: List[str], machine_lanes: Dict[str, List[List[dict]]], out_xlsx: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "排产"

    ws.column_dimensions["A"].width = MACHINE_COL_WIDTH
    for i in range(len(dates)):
        ws.column_dimensions[get_column_letter(2 + i)].width = DATE_COL_WIDTH

    grid_side = Side(style=GRID_STYLE, color=GRID_COLOR)
    grid_border = Border(left=grid_side, right=grid_side, top=grid_side, bottom=grid_side)

    sep_side = Side(style=MACHINE_SEP_STYLE, color=MACHINE_SEP_COLOR)

    ws["A1"] = TITLE
    ws["A1"].font = Font(bold=True, size=18)

    ws.row_dimensions[2].height = HEADER_ROW_HEIGHT
    ws.row_dimensions[3].height = HEADER_ROW_HEIGHT

    ws["A2"] = "机台"
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].font = Font(bold=True, size=12)
    ws.merge_cells("A2:A3")  # 合并后不要再写 A3

    for idx, d in enumerate(dates):
        c = 2 + idx
        ws.cell(2, c, f"{d.month}/{d.day}").alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(3, c, weekday_cn(d)).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(2, c).font = Font(bold=True, size=12)
        ws.cell(3, c).font = Font(bold=True, size=12)

    row = 4
    last_col = 1 + len(dates)

    for m in machines:
        lanes = machine_lanes[m]
        lane_count = len(lanes)
        start_row = row
        end_row = row + lane_count - 1

        ws.cell(start_row, 1, m).font = Font(bold=True, size=12)
        ws.cell(start_row, 1).alignment = Alignment(horizontal="left", vertical="center")
        if lane_count > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)

        for r in range(start_row, end_row + 1):
            ws.row_dimensions[r].height = BASE_LANE_ROW_HEIGHT
            for c in range(1, last_col + 1):
                cell = ws.cell(r, c)
                cell.border = grid_border
                if c != 1:
                    cell.value = ""
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        # ✅ 机台分隔线（上/下边框加粗）
        if APPLY_MACHINE_SEPARATORS:
            if MACHINE_SEP_TOP:
                for c in range(1, last_col + 1):
                    apply_border_keep(ws.cell(start_row, c), top=sep_side)
            if MACHINE_SEP_BOTTOM:
                for c in range(1, last_col + 1):
                    apply_border_keep(ws.cell(end_row, c), bottom=sep_side)

        row = end_row + 1

    ws.freeze_panes = "B4"
    wb.save(out_xlsx)


def build_layout_json(
    dates: List[Any],
    machines: List[str],
    machine_lanes: Dict[str, List[List[dict]]],
    win_start,
    win_end,
    out_layout: str
):
    base_row = 4
    date_base_col = 2
    machine_start_row: Dict[str, int] = {}
    r = base_row
    for m in machines:
        machine_start_row[m] = r
        r += len(machine_lanes[m])

    items = []
    for m in machines:
        start_r = machine_start_row[m]
        for lane_idx, lane_tasks in enumerate(machine_lanes[m]):
            row_num = start_r + lane_idx
            lane_tasks_sorted = sorted(lane_tasks, key=lambda x: (x["start_ms"], x["end_ms"], x["text"]))
            for t in lane_tasks_sorted:
                s_eff, e_eff = clip_interval(t, win_start, win_end)
                if s_eff is None:
                    continue
                start_c = date_base_col + (s_eff - win_start).days
                end_c = date_base_col + (e_eff - win_start).days
                items.append({
                    "task_id": t["task_id"],
                    "machine": m,
                    "lane": lane_idx,
                    "row": row_num,
                    "col_start": start_c,
                    "col_end": end_c,
                    "start_date": s_eff.isoformat(),
                    "end_date": e_eff.isoformat(),
                    "start_ms": t["start_ms"],
                    "end_ms": t["end_ms"],
                    "text": t["text"],
                    "rgb": None if t["rgb"] is None else [t["rgb"][0], t["rgb"][1], t["rgb"][2]],
                })

    payload = {
        "title": TITLE,
        "win_start": win_start.isoformat(),
        "win_end": win_end.isoformat(),
        "dates": [d.isoformat() for d in dates],
        "machines": machines,
        "items": items,
    }
    with open(out_layout, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def main():
    ensure_dir(OUTPUT_DIR)

    records_path = RECORDS_JSON
    if AUTO_LATEST:
        records_path = pick_latest_records(OUTPUT_DIR)

    with open(records_path, "r", encoding="utf-8") as f:
        records_doc = json.load(f)

    tasks_all = extract_tasks(records_doc)
    if not tasks_all:
        raise RuntimeError("未抽取到任务：请确认 records.raw.json 中包含 机台/开始日期/结束日期/甘特图文本")

    win_start, win_end = compute_window(tasks_all)
    dates = build_dates(win_start, win_end)

    tasks_in_win = [t for t in tasks_all if not (t["end"] < win_start or t["start"] > win_end)]
    machines = build_machine_list(tasks_in_win)
    if not machines:
        raise RuntimeError("窗口内无机台任务：请调整日期范围")

    tasks_by_machine: Dict[str, List[dict]] = {}
    for t in tasks_in_win:
        tasks_by_machine.setdefault(t["machine"], []).append(t)

    machine_lanes: Dict[str, List[List[dict]]] = {}
    for m in machines:
        machine_lanes[m] = assign_lanes(tasks_by_machine.get(m, []), win_start, win_end) or [[]]

    print_preview(tasks_in_win, win_start, win_end)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_xlsx = os.path.join(OUTPUT_DIR, f"排产_框架_{ts}.xlsx")
    out_layout = os.path.join(OUTPUT_DIR, f"排产_layout_{ts}.json")
    out_csv = os.path.join(OUTPUT_DIR, f"排产_核对_{ts}.csv")

    write_check_csv(tasks_in_win, out_csv)
    build_frame_xlsx(dates, machines, machine_lanes, out_xlsx)
    build_layout_json(dates, machines, machine_lanes, win_start, win_end, out_layout)

    print(f"OK: {out_xlsx}")
    print(f"layout: {out_layout}")
    print(f"check: {out_csv}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)
