"""
程序简介：计算采购价格和成本，输出可核对结果。
主要逻辑：读取所需配置或输入数据，执行本文件负责的处理步骤，并把结果写入本地文件或输出到命令行。
配置说明：涉及企微或飞书凭证时，优先读取 PEIFANG_ENV_PROFILE、WECOM_ENV_PROFILE、FEISHU_ENV_PROFILE 选择公司配置档案；未设置时兼容原来的 .env 变量。
"""

# -*- coding: utf-8 -*-
import sys
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ================================
# 📂 配置
# ================================
DESKTOP = Path(r"C:\Users\ishel\Desktop\坚果备份\A四川和裕达新材料有限公司\24品质技术开发\运算文件夹")
MATRIX_FILE_GLOB = "物料清单合并_*.xlsx"
MATRIX_SHEET_NAME = "父子件比例矩阵"

PRICE_FILE_GLOB = "采购价格波动分析表*.xlsx"
PRICE_SHEET_NAME = "第一页"
PRICE_HEADER_ROW = 4
COL_CODE_NAME   = "存货编码"
COL_PRICE_NAME  = "最近价格1"
COL_ITEM_NAME   = "存货名称"
COL_SUPP_NAME   = "供应商"

ROW_TO_WRITE_PRICE = 3
NUM_FMT_PRICE = "0.00"
ROW_HEADER = 4
EXPECTED_HEADERS = ["版本", "编码", "名称", "利润率", "利润", "售价", "成本"]

# ================================
# 🔧 日志与包装
# ================================
def log(msg: str):
    print(msg if len(msg) <= 300 else (msg[:297] + "..."))

def step(title, fn, *args, **kwargs):
    log(f"▶ 开始：{title}")
    try:
        result, info = fn(*args, **kwargs)
        brief = "; ".join(f"{k}={v}" for k, v in info.items()) if isinstance(info, dict) else str(info)
        log(f"✅ 成功：{title} ｜ {brief}")
        return result
    except Exception as e:
        log(f"❌ 失败：{title} ｜ {type(e).__name__}: {e}")
        raise

# ================================
# 🔧 工具函数
# ================================
def s(v):
    return "" if v is None else str(v).strip()

def to_float(x):
    try:
        v = pd.to_numeric(x)
        return float(v) if pd.notna(v) else None
    except Exception:
        return None

def normalize_unit(u: str) -> str:
    u = s(u).lower()
    if u in {"kg", "公斤", "千克"}: return "kg"
    if u in {"g", "克"}: return "g"
    return "unknown"

def find_latest(glob_pat: str):
    files = sorted(DESKTOP.glob(glob_pat), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"未找到文件：{glob_pat}")
    path = files[0]
    show = "; ".join(f.name for f in files[:3])
    return path, {"latest": path.name, "candidates(top3)": show}

def open_matrix():
    path, _ = find_latest(MATRIX_FILE_GLOB)
    wb = load_workbook(path)
    if MATRIX_SHEET_NAME not in wb.sheetnames:
        raise KeyError(f"缺少工作表：{MATRIX_SHEET_NAME}")
    ws = wb[MATRIX_SHEET_NAME]
    return (wb, ws, path), {"file": path.name, "sheet": MATRIX_SHEET_NAME}

def map_row_headers(ws, row_idx: int):
    header = {}
    for c in range(1, ws.max_column + 1):
        key = s(ws.cell(row=row_idx, column=c).value)
        if key:
            header[key] = c
    missing = [h for h in EXPECTED_HEADERS if h not in header]
    if missing:
        raise KeyError(f"第{row_idx}行缺少表头: {missing}")
    return header, {"cols": len(header), "cost_col": get_column_letter(header['成本'])}

def find_data_start_col(ws, header_cols_max: int):
    start = header_cols_max + 1
    for c in range(start, ws.max_column + 1):
        if s(ws.cell(row=1, column=c).value):
            return c, {"start_col": get_column_letter(c)}
    return start, {"start_col": get_column_letter(start)}

def build_code_to_col(ws, data_start_col: int):
    code2col = {}
    for c in range(data_start_col, ws.max_column + 1):
        key = s(ws.cell(row=1, column=c).value)
        if key:
            code2col[key] = c
    return code2col, {"code_cols": len(code2col)}

def copy_price_sheet(wb):
    price_path, info = find_latest(PRICE_FILE_GLOB)
    src_wb = load_workbook(price_path, data_only=True)
    if PRICE_SHEET_NAME not in src_wb.sheetnames:
        raise KeyError(f"价格簿缺少工作表：{PRICE_SHEET_NAME}")
    src_ws = src_wb[PRICE_SHEET_NAME]
    copied_name = PRICE_SHEET_NAME
    k = 1
    while copied_name in wb.sheetnames:
        k += 1
        copied_name = f"{PRICE_SHEET_NAME}({k})"
    dst_ws = wb.create_sheet(title=copied_name)
    for r in src_ws.iter_rows(values_only=True):
        dst_ws.append(list(r))
    return (copied_name, price_path), {"copied_sheet": copied_name, "price_file": price_path.name}

# ================================
# 🔧 单位换算（仅 g→kg；其余不换算）
# ================================
def to_price_per_kg(price_raw, unit_str):
    v = to_float(price_raw)
    if v is None:
        return None, "skip:none"
    u = normalize_unit(unit_str)
    if u == "kg":
        return v, "kg"
    if u == "g":
        return v * 1000.0, "g→kg"
    return None, "skip:unit"  # 异常单位：不换算，跳过

def parse_and_write_prices(wb, ws, copied_name, code2col):
    hws = wb[copied_name]
    header_price = {s(hws.cell(row=PRICE_HEADER_ROW, column=c).value): c for c in range(1, hws.max_column + 1)}
    col_code = header_price.get(COL_CODE_NAME)
    col_price = header_price.get(COL_PRICE_NAME)
    col_name = header_price.get(COL_ITEM_NAME)
    col_supp = header_price.get(COL_SUPP_NAME)
    if not col_code or not col_price:
        raise KeyError(f"价格表缺少列：{COL_CODE_NAME}/{COL_PRICE_NAME}")
    unit_col = col_price - 1  # 价格左侧单元格为单位

    buckets = {}
    # 异常单位收集：[(name, unit, in_matrix_bool)]
    abnormal_entries = []

    for r in range(PRICE_HEADER_ROW + 1, hws.max_row + 1):
        code = s(hws.cell(row=r, column=col_code).value)
        unit = hws.cell(row=r, column=unit_col).value
        name = s(hws.cell(row=r, column=col_name).value) if col_name else "—"
        supp = s(hws.cell(row=r, column=col_supp).value) if col_supp else "—"
        price_raw = hws.cell(row=r, column=col_price).value

        norm_u = normalize_unit(unit)
        if norm_u == "unknown":
            # 仅打印名称+单位，不换算，不入桶
            abnormal_entries.append((name, s(unit), (code in code2col)))
            continue

        price_kg, tag = to_price_per_kg(price_raw, unit)

        # 单位为 g：仅当矩阵首行存在该编码才打印换算明细
        if norm_u == "g" and code in code2col:
            pr = to_float(price_raw)
            if pr is not None:
                log(f"↻ 单位换算｜编码={code}｜名称={name}｜供应商={supp}｜{pr}/g → {pr*1000.0}/kg")

        if not code or price_kg is None:
            continue

        buckets.setdefault(code, []).append({
            "price_raw": to_float(price_raw),
            "unit": s(unit),
            "price": price_kg,
            "name": name,
            "supp": supp
        })

    # 异常单位打印（只打印“名称+单位”与总计、其中可写入矩阵的数量）
    if abnormal_entries:
        in_matrix_cnt = sum(1 for _, _, ok in abnormal_entries if ok)
        log(f"⚠️ 异常单位（非 kg/g）：{len(abnormal_entries)} 条（其中在矩阵首行存在：{in_matrix_cnt} 条）")
        for nm, un, _ok in abnormal_entries:
            log(f"   异常单位｜名称={nm}｜单位={un}")

    # 写入与多价打印（kg不展示换算；g标注参与换算）
    filled = 0
    multi_choice_cnt = 0
    direct_write_kg_unique = 0  # 唯一且kg的直接写入计数

    for code, items in buckets.items():
        if not items:
            continue

        # 唯一且kg：只计数与写入，不逐条打印候选
        if len(items) == 1 and normalize_unit(items[0]["unit"]) == "kg" and (code in code2col):
            best = items[0]
            col = code2col.get(code)
            if col:
                cell = ws.cell(row=ROW_TO_WRITE_PRICE, column=col, value=best["price"])
                cell.number_format = NUM_FMT_PRICE
                filled += 1
                direct_write_kg_unique += 1
            continue

        prices = [it["price"] for it in items]
        if len(set(prices)) > 1:
            multi_choice_cnt += 1
            log(f"ℹ️ 多价选择：{code}")
            for it in items:
                u = normalize_unit(it["unit"])
                if u == "kg":
                    log(f"   候选｜编码={code}｜名称={it['name']}｜供应商={it['supp']}｜原价={it['price_raw']}／kg")
                elif u == "g":
                    log(f"   候选｜编码={code}｜名称={it['name']}｜供应商={it['supp']}｜原价={it['price_raw']}／g 〔⚠️ g→kg 换算参与比较〕")
                else:
                    log(f"   候选｜编码={code}｜名称={it['name']}｜供应商={it['supp']}｜原价={it['price_raw']}／{it['unit']} 〔⚠ 异常单位，未参与比较〕")

        # 选择最小元/kg并写入
        best = min(items, key=lambda z: z["price"])
        col = code2col.get(code)
        if col:
            cell = ws.cell(row=ROW_TO_WRITE_PRICE, column=col, value=best["price"])
            cell.number_format = NUM_FMT_PRICE
            filled += 1
            # 仅在不是“唯一且kg”的情况下，给一条简短结果
            if not (len(items) == 1 and normalize_unit(items[0]["unit"]) == "kg"):
                log(f"   → 取最小｜编码={code}｜供应商={best['supp']}｜折合(元/kg)={best['price']}")

    # 唯一且kg的直接写入数量汇总
    log(f"✅ 直接写入（唯一且kg）：{direct_write_kg_unique} 条")

    return filled, {
        "codes_found": len(buckets),
        "written": filled,
        "multi_choice": multi_choice_cnt,
        "direct_unique_kg": direct_write_kg_unique
    }

def calc_row_costs(ws, data_start_row: int, data_start_col: int, cost_col: int, price_row: int, num_fmt_cost: str):
    valid_cols = [c for c in range(data_start_col, ws.max_column + 1) if s(ws.cell(row=1, column=c).value)]
    filled = 0
    for r in range(data_start_row, ws.max_row + 1):
        total = 0.0; used = False
        for c in range(min(valid_cols) if valid_cols else data_start_col, (max(valid_cols) if valid_cols else data_start_col) + 1):
            if c not in valid_cols:
                continue
            pct = to_float(ws.cell(row=r, column=c).value)
            price = to_float(ws.cell(row=price_row, column=c).value)
            if pct is not None and price is not None:
                total += pct * price; used = True
        if used:
            cell = ws.cell(row=r, column=cost_col, value=total)
            cell.number_format = num_fmt_cost
            filled += 1
    return filled, {"rows_calc": filled, "cost_col": get_column_letter(cost_col)}

def save_wb(wb, path: Path):
    try:
        wb.save(path)
        return None, {"saved": str(path)}
    except PermissionError:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = path.with_name(path.stem + f"_calc_{ts}" + path.suffix)
        wb.save(alt)
        return None, {"saved_alt": str(alt), "reason": "file locked"}

# ================================
# 🚀 主流程
# ================================
def main():
    (wb, ws, matrix_path) = step("打开矩阵文件", open_matrix)
    hdr = step("读取表头", map_row_headers, ws, ROW_HEADER)
    header_cols_max = max(hdr[h] for h in EXPECTED_HEADERS)
    cost_col = hdr["成本"]

    data_start_col = step("定位数据起始列", find_data_start_col, ws, header_cols_max)
    code2col = step("建立子件编码→列号映射", build_code_to_col, ws, data_start_col)
    data_start_row = ROW_HEADER + 1

    try:
        (copied_name, price_path) = step("复制价格表《第一页》", copy_price_sheet, wb)
        _ = step("解析价格并写入第3行", parse_and_write_prices, wb, ws, copied_name, code2col)
    except Exception as e:
        log(f"⚠️ 跳过价格写入：{e}")

    _ = step("计算成本", calc_row_costs, ws, data_start_row, data_start_col, cost_col, ROW_TO_WRITE_PRICE, NUM_FMT_PRICE)
    _ = step("保存文件", save_wb, wb, matrix_path)
    log("🏁 全流程完成。")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"❌ 终止：{e}")
        sys.exit(1)
