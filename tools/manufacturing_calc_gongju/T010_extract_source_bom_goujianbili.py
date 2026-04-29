"""
程序简介：从原始数据提取 BOM 或构件比例，生成后续矩阵处理输入。
主要逻辑：读取所需配置或输入数据，执行本文件负责的处理步骤，并把结果写入本地文件或输出到命令行。
配置说明：涉及企微或飞书凭证时，优先读取 PEIFANG_ENV_PROFILE、WECOM_ENV_PROFILE、FEISHU_ENV_PROFILE 选择公司配置档案；未设置时兼容原来的 .env 变量。
"""

import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ================================
# 📂 1. 配置路径和参数
# ================================
INPUT_PATH = r"C:\Users\ishel\Desktop\坚果备份\A四川和裕达新材料有限公司\24品质技术开发\运算文件夹\物料清单.xlsx"
OUTPUT_FOLDER = r"C:\Users\ishel\Desktop\坚果备份\A四川和裕达新材料有限公司\24品质技术开发\运算文件夹"
SHEET_MATERIAL = "物料清单"
SHEET_COMPONENT = "子件明细"
SHEET_RESULT = "父件子件明细"

CUSTOM_COLUMNS_ORDER = [
    "版本号_子件", "父件编码", "子件编码", "父件名称", "子件名称",
    "规格型号_子件", "计量单位_子件", "需用数量", "生产数量_子件", "规格型号_父件",
    "创建时间", "标准用量", "版本号_父件", "存货图片", "备注", "子件BOM",
    "子件默认BOM", "损耗率%", "预出仓库编码", "预出仓库", "倒冲料",
    "材料倒冲方式", "计量单位_父件", "生产数量_父件", "生产车间编码", "生产车间",
    "预入仓库编码", "预入仓库", "默认BOM", "成品率%", "停用"
]
SKIP_CHILD_CODES = {"30008", "30004", "90011", "30009", "90024", "03000012", "3000011", "0300001"}
GRAM_UNIT = "克"

# ================================
# 📥 2. 读取数据
# ================================
def load_frames(input_path, sheet_material, sheet_component):
    print("📥 Step2: 读取数据中…")
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"未找到输入文件：{input_path}")
    df_m = pd.read_excel(input_path, sheet_name=sheet_material)
    df_c = pd.read_excel(input_path, sheet_name=sheet_component)
    print(f"   物料清单行:{len(df_m)}, 子件明细行:{len(df_c)}")
    return df_m, df_c

# ================================
# 🔗 3. 合并数据
# ================================
def merge_frames(df_c, df_m):
    print("🔗 Step3: 合并父子件数据…")
    merged = pd.merge(df_c, df_m, on="父件编码", how="left",
                      suffixes=('_子件', '_父件'))
    print(f"   合并后行数:{len(merged)} 列数:{len(merged.columns)}")
    return merged

# ================================
# 🧾 4. 字段校验与排序
# ================================
def align_columns(merged, desired):
    print("🧾 Step4: 校验并排序字段…")
    merged_cols = list(merged.columns)
    missing = [c for c in desired if c not in merged_cols]
    extra = [c for c in merged_cols if c not in desired]
    aligned = merged.reindex(columns=[c for c in desired if c in merged.columns])
    print(f"   缺列:{missing if missing else '无'} | 多列:{extra if extra else '无'}")
    return aligned, missing, extra

# ================================
# 🔁 5. 去重（在计算比例之前执行）
# ================================
def dedup_items(df):
    print("🔁 Step5: 去重父件子件明细…（先去重，再算比例）")
    keys = ["版本号_子件", "父件编码", "子件编码"]
    for k in keys:
        if k not in df.columns:
            print(f"   ⚠️ 缺少关键列 {k}，跳过去重。")
            return df

    # 规范化临时键值以提升匹配稳定性，不改动原数据
    tmp = df[keys].astype(str).apply(lambda s: s.str.strip())
    total = len(df)
    dup_all_mask = tmp.duplicated(subset=keys, keep=False)
    dup_drop_mask = tmp.duplicated(subset=keys, keep="first")

    groups_cnt = int(tmp.loc[dup_all_mask].drop_duplicates(subset=keys).shape[0])
    removed = int(dup_drop_mask.sum())
    kept = total - removed

    # 打印被删除的行（尽量多给信息，但列缺失也不报错）
    if removed > 0:
        extra_cols = [c for c in ["父件名称", "子件名称", "规格型号_子件", "需用数量"] if c in df.columns]
        print("   ⚠️ 以下重复行被删除：")
        print(df.loc[dup_drop_mask, keys + extra_cols].to_string(index=False))

    df_dedup = df.loc[~dup_drop_mask].copy()
    df_dedup = df_dedup[df.columns]  # 保持原列顺序

    print(f"   重复组:{groups_cnt} | 删除行:{removed} | 保留行:{kept} | 原始行:{total}")
    return df_dedup

# ================================
# ➕ 6. 计算比例
# ================================
def compute_ratio_series(df):
    print("➕ Step6: 计算比例列…")
    qty = pd.to_numeric(df.get("需用数量"), errors="coerce")
    unit = df.get("计量单位_子件").astype(str).str.strip() if "计量单位_子件" in df.columns else ""
    qty_kg = qty.where(unit != GRAM_UNIT, qty / 1000)
    child = df.get("子件编码").astype(str).str.strip() if "子件编码" in df.columns else ""
    mask = qty_kg.notna() & (~child.isin(SKIP_CHILD_CODES))

    grp = (
        df.get("版本号_子件").astype(str).str.strip().fillna("")
        .str.cat(df.get("父件编码").astype(str).str.strip().fillna(""), sep="||")
    )
    denom = qty_kg.where(mask, 0).groupby(grp).transform("sum")
    ratio = (qty_kg / denom).where(mask & (denom > 0))
    print(f"   比例非空:{ratio.notna().sum()}")
    return ratio

def insert_ratio_column(df):
    ratio = compute_ratio_series(df)
    df = df.copy()
    df["比例"] = ratio
    if "生产数量_子件" in df.columns:
        pos = df.columns.get_loc("生产数量_子件") + 1
        cols = list(df.columns)
        cols.insert(pos, cols.pop(cols.index("比例")))
        df = df[cols]
    return df

# ================================
# 💾 7. 保存
# ================================
def save_as_excel(path, df_m, df_c, merged, s1, s2, s3, df_items):
    print("💾 Step7: 保存Excel…")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_m.to_excel(writer, sheet_name=s1, index=False)
        df_c.to_excel(writer, sheet_name=s2, index=False)
        merged.to_excel(writer, sheet_name=s3, index=False)
        df_items.to_excel(writer, sheet_name="物料编码及名称", index=False)
    print(f"   已保存: {path}")

# ================================
# 🎨 8. 格式
# ================================
def format_result_sheet(path, sheet_result):
    print("🎨 Step8: 设置格式…")
    wb = load_workbook(path)
    ws = wb[sheet_result]
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="left", vertical="center")
    wb.save(path)
    print("   格式化完成")

# ================================
# 🚀 9. 主流程
# ================================
def main():
    print("🚀 主流程开始")
    df_m, df_c = load_frames(INPUT_PATH, SHEET_MATERIAL, SHEET_COMPONENT)
    merged = merge_frames(df_c, df_m)
    aligned, missing, extra = align_columns(merged, CUSTOM_COLUMNS_ORDER)

    # 先去重，再算比例
    aligned = dedup_items(aligned)
    aligned = insert_ratio_column(aligned)

    from copy import deepcopy
    df_items = deepcopy(df_c[["子件编码", "子件名称"]]) if "子件编码" in df_c else pd.DataFrame()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUT_FOLDER, f"物料清单合并_{ts}.xlsx")
    save_as_excel(output_path, df_m, df_c, aligned,
                  SHEET_MATERIAL, SHEET_COMPONENT, SHEET_RESULT, df_items)
    format_result_sheet(output_path, SHEET_RESULT)

    print("✅ 流程结束")

if __name__ == "__main__":
    main()
