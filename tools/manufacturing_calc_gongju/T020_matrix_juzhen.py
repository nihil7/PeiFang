"""
程序简介：基于提取结果构建或整理矩阵数据。
主要逻辑：读取所需配置或输入数据，执行本文件负责的处理步骤，并把结果写入本地文件或输出到命令行。
配置说明：涉及企微或飞书凭证时，优先读取 PEIFANG_ENV_PROFILE、WECOM_ENV_PROFILE、FEISHU_ENV_PROFILE 选择公司配置档案；未设置时兼容原来的 .env 变量。
"""

# build_matrix.py
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# ================================
# 📂 配置路径 / 工作表名（集中）
# ================================
DESKTOP = Path(r"C:\Users\ishel\Desktop\坚果备份\A四川和裕达新材料有限公司\24品质技术开发\运算文件夹")
INPUT_GLOB = "物料清单合并_*.xlsx"
SHEET_SOURCE = "父件子件明细"
SHEET_MATRIX = "父子件比例矩阵"

# ================================
# ⚙️ 参数
# ================================
AGG_FUNC = "sum"   # sum/mean/first
SUM_TOL = 1e-6     # 行求和≈100%容差

# ================================
# 🔧 基础打印/步骤包装
# ================================
def log(msg: str):  # 单次≤300字
    print(msg if len(msg) <= 300 else (msg[:297] + "..."))

def run_step(title, fn, *args, **kwargs):
    log(f"▶ 开始：{title}")
    try:
        result, info = fn(*args, **kwargs)
        # 将关键信息拼成短句打印
        if isinstance(info, dict):
            brief = "; ".join(f"{k}={v}" for k, v in info.items())
        else:
            brief = str(info)
        log(f"✅ 成功：{title} ｜ {brief}")
        return result
    except Exception as e:
        log(f"❌ 失败：{title} ｜ {type(e).__name__}: {e}")
        raise

# ================================
# 🔧 工具函数
# ================================
def find_latest(glob_pat: str):
    files = sorted(DESKTOP.glob(glob_pat), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        raise FileNotFoundError(f"未找到文件：{glob_pat}")
    path = files[0]
    return path, {"file": path.name}

def _norm(s):
    if pd.isna(s): return ""
    return str(s).strip()

# 1) 读取与聚合
def read_and_aggregate(input_path: Path):
    usecols = ["版本号_子件","父件编码","父件名称","子件编码","子件名称","比例"]
    df = pd.read_excel(input_path, sheet_name=SHEET_SOURCE, dtype=str, usecols=usecols)
    for c in ["版本号_子件","父件编码","父件名称","子件编码","子件名称"]:
        df[c] = df[c].map(_norm)
    df["比例"] = pd.to_numeric(df["比例"], errors="coerce")
    df = df[(df["版本号_子件"]!="") & (df["父件编码"]!="") & (df["子件编码"]!="")]
    if df.empty:
        raise RuntimeError("清洗后数据为空。")

    key_cols = ["版本号_子件","父件编码","父件名称","子件编码","子件名称"]
    has_dup = df.duplicated(subset=key_cols, keep=False).any()
    if AGG_FUNC == "sum":
        agg_df = df.groupby(key_cols, as_index=False, dropna=False)["比例"].sum()
    elif AGG_FUNC == "mean":
        agg_df = df.groupby(key_cols, as_index=False, dropna=False)["比例"].mean()
    elif AGG_FUNC == "first":
        agg_df = df.sort_index().groupby(key_cols, as_index=False, dropna=False)["比例"].first()
    else:
        raise ValueError("AGG_FUNC 仅支持 sum/mean/first")

    dup_sample = 0
    if has_dup:
        dups = (df.groupby(key_cols)
                  .filter(lambda x: len(x)>1)
                  .groupby(key_cols)["比例"].apply(list)
                  .reset_index())
        dup_sample = min(8, len(dups))

    info = {
        "rows_raw": len(df),
        "rows_agg": len(agg_df),
        "dup_groups_sample": dup_sample,
        "agg": AGG_FUNC
    }
    return agg_df, info

# 2) 透视为矩阵
def make_pivot(agg_df: pd.DataFrame):
    father_order = agg_df.drop_duplicates(["版本号_子件","父件编码","父件名称"])[
        ["版本号_子件","父件编码","父件名称"]
    ]
    child_order  = agg_df.drop_duplicates(["子件编码","子件名称"])[
        ["子件编码","子件名称"]
    ]
    mat = pd.pivot_table(
        agg_df,
        index=["版本号_子件","父件编码","父件名称"],
        columns=["子件编码","子件名称"],
        values="比例",
        aggfunc="first"
    )
    mat = mat.reindex(pd.MultiIndex.from_frame(father_order), axis=0)\
             .reindex(pd.MultiIndex.from_frame(child_order), axis=1)
    info = {"father_rows": mat.shape[0], "child_cols": mat.shape[1]}
    return mat, info

# 3) 写入工作簿（仅矩阵与基础格式）
def write_matrix(input_path: Path, mat: pd.DataFrame):
    wb = load_workbook(input_path)
    if SHEET_MATRIX in wb.sheetnames:
        del wb[SHEET_MATRIX]
    ws = wb.create_sheet(title=SHEET_MATRIX)

    # 子件表头：1行=子件编码，2行=子件名称；数据起点 D3
    start_row, start_col = 3, 4
    for j, (ccode, cname) in enumerate(mat.columns, start=start_col):
        ws.cell(row=1, column=j, value=ccode).alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=j, value=cname).alignment = Alignment(horizontal="center")

    # 左侧三列
    for i, (vcode, fcode, fname) in enumerate(mat.index, start=start_row):
        ws.cell(row=i, column=1, value=vcode)
        ws.cell(row=i, column=2, value=fcode)
        ws.cell(row=i, column=3, value=fname)

    # 数据区
    n_rows, n_cols = mat.shape
    for i in range(n_rows):
        for j in range(n_cols):
            v = mat.iat[i, j]
            if pd.notna(v):
                ws.cell(row=start_row+i, column=start_col+j, value=float(v))

    # 百分比格式
    for r in range(start_row, start_row+n_rows):
        for c in range(start_col, start_col+n_cols):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0000%"

    # 列宽与抬头居中
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    for c in range(1, ws.max_column+1):
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=c).alignment = Alignment(horizontal="center")

    wb.save(input_path)
    info = {"sheet": SHEET_MATRIX, "saved": input_path.name}
    return input_path, info

# 4) 校验行和≈100%
def check_sum_approx_100(input_path: Path):
    wb = load_workbook(input_path, data_only=True)
    ws = wb[SHEET_MATRIX]
    # 数据区从 D3 开始；此处只统计非空数据的和
    start_row, start_col = 3, 4
    max_row, max_col = ws.max_row, ws.max_column

    bad = []
    # 计算有效数据行数 = 从 start_row 到 max_row，直到这一行在数据区全空则可视为尾部空行（简化：遍历全表）
    for r in range(start_row, max_row+1):
        s = 0.0
        has_num = False
        for c in range(start_col, max_col+1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)):
                s += float(v)
                has_num = True
        if has_num and not (abs(s - 1.0) <= SUM_TOL):
            vcode = ws.cell(row=r, column=1).value
            fcode = ws.cell(row=r, column=2).value
            bad.append(f"{vcode}/{fcode}={s:.6f}")

    info = {"rows_checked": max(0, max_row - start_row + 1),
            "bad_rows": len(bad),
            "sample": "; ".join(bad[:3]) + (f"...(+{len(bad)-3})" if len(bad) > 3 else "")}
    return len(bad), info

# ================================
# 🚀 主流程
# ================================
def main():
    input_path = run_step("定位最新输入文件", find_latest, INPUT_GLOB)
    agg_df     = run_step("读取与聚合", read_and_aggregate, input_path)
    mat        = run_step("生成透视矩阵", make_pivot, agg_df)
    _          = run_step("写入矩阵与基础格式", write_matrix, input_path, mat)
    _          = run_step("校验行和≈100%", check_sum_approx_100, input_path)
    log("🏁 程序一完成。")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"❌ 终止：{type(e).__name__}: {e}")
        sys.exit(1)
