import os
import glob
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =========================
# 配置区（只改这里）
# =========================
INPUT_DIR = r"C:\Users\ishel\Desktop"   # 你要扫描的文件夹
KEYWORD = "色差"                        # 文件名包含该关键词才处理
OUTPUT_SUFFIX = "_RGB处理"              # 输出文件后缀

SHEET_NAME = None                       # 指定工作表名；None=默认用活动表
START_ROW = 2                           # 数据起始行（表头在第1行）
COL_L = "F"                             # L* 列
COL_A = "G"                             # a* 列
COL_B = "H"                             # b* 列
COL_RGB_TEXT = "I"                      # 写入 rgb(r,g,b) 的列
COL_RGB_CELL = "J"                      # 填充颜色的列
WRITE_RGB_TEXT_IN_J = False             # True: J列也写rgb文本；False: 只填充颜色
# =========================


def lab_to_rgb(L, a, b):
    """
    CIE L*a*b* (D65) -> sRGB 0..255
    """
    # Reference white D65
    Xn, Yn, Zn = 95.047, 100.000, 108.883

    fy = (L + 16.0) / 116.0
    fx = fy + (a / 500.0)
    fz = fy - (b / 200.0)

    def f_inv(t):
        if t ** 3 > 0.008856:
            return t ** 3
        return (t - 16.0 / 116.0) / 7.787

    X = Xn * f_inv(fx)
    Y = Yn * f_inv(fy)
    Z = Zn * f_inv(fz)

    # XYZ (0..1)
    x = X / 100.0
    y = Y / 100.0
    z = Z / 100.0

    # XYZ -> linear RGB (sRGB D65)
    r_lin =  3.2406 * x + (-1.5372) * y + (-0.4986) * z
    g_lin = -0.9689 * x +  1.8758 * y +  0.0415 * z
    b_lin =  0.0557 * x + (-0.2040) * y +  1.0570 * z

    def gamma(u):
        # clip first
        u = 0.0 if u < 0.0 else (1.0 if u > 1.0 else u)
        if u <= 0.0031308:
            return 12.92 * u
        return 1.055 * (u ** (1.0 / 2.4)) - 0.055

    r = int(round(255.0 * gamma(r_lin)))
    g = int(round(255.0 * gamma(g_lin)))
    bb = int(round(255.0 * gamma(b_lin)))

    # ensure 0..255
    r = 0 if r < 0 else (255 if r > 255 else r)
    g = 0 if g < 0 else (255 if g > 255 else g)
    bb = 0 if bb < 0 else (255 if bb > 255 else bb)
    return r, g, bb


def safe_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def process_file(xlsx_path: str):
    wb = load_workbook(xlsx_path)
    ws = wb[SHEET_NAME] if (SHEET_NAME and SHEET_NAME in wb.sheetnames) else wb.active

    # headers
    ws[f"{COL_RGB_TEXT}1"].value = "转化为RGB"
    ws[f"{COL_RGB_CELL}1"].value = "RGB显示效果"

    max_row = ws.max_row
    filled = 0

    for r in range(START_ROW, max_row + 1):
        L = safe_float(ws[f"{COL_L}{r}"].value)
        a = safe_float(ws[f"{COL_A}{r}"].value)
        b = safe_float(ws[f"{COL_B}{r}"].value)

        # 如果这一行三个都空，就跳过
        if L is None and a is None and b is None:
            continue

        # 缺一个就不算（你也可以改成缺失按0）
        if L is None or a is None or b is None:
            ws[f"{COL_RGB_TEXT}{r}"].value = ""
            ws[f"{COL_RGB_CELL}{r}"].value = ""
            ws[f"{COL_RGB_CELL}{r}"].fill = PatternFill(fill_type=None)
            continue

        rr, gg, bb = lab_to_rgb(L, a, b)
        rgb_text = f"rgb({rr},{gg},{bb})"
        ws[f"{COL_RGB_TEXT}{r}"].value = rgb_text

        hex_color = f"{rr:02X}{gg:02X}{bb:02X}"
        ws[f"{COL_RGB_CELL}{r}"].fill = PatternFill(
            start_color=hex_color, end_color=hex_color, fill_type="solid"
        )
        ws[f"{COL_RGB_CELL}{r}"].value = rgb_text if WRITE_RGB_TEXT_IN_J else ""

        filled += 1

    src = Path(xlsx_path)
    out_path = src.with_name(src.stem + OUTPUT_SUFFIX + src.suffix)
    wb.save(out_path)
    return str(out_path), filled


def main():
    pattern = os.path.join(INPUT_DIR, f"*{KEYWORD}*.xls*")
    files = [f for f in glob.glob(pattern) if not os.path.basename(f).startswith("~$")]

    if not files:
        print("未找到含“色差”的Excel文件。")
        return

    for f in files:
        try:
            out, n = process_file(f)
            print(f"OK: {os.path.basename(f)} -> {os.path.basename(out)} 行数={n}")
        except Exception as e:
            print(f"FAIL: {os.path.basename(f)} 错误={e}")

    print("完成。")


if __name__ == "__main__":
    main()
